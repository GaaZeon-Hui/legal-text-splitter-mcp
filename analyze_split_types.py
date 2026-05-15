"""
原文拆分类型计数 & 逻辑递增组判断脚本

用法:
  python analyze_split_types.py                    内置测试文本
  python analyze_split_types.py <文件路径>           分析单个文件
  python analyze_split_types.py --id <law_id>       从数据库拉取单条分析
  python analyze_split_types.py --batch             从 Excel 批量读取 law_id 并分析输出
"""
import sys
import re
import os
try:
    import openpyxl
except ImportError:
    pass
try:
    import pymysql
except ImportError:
    pass
from collections import OrderedDict

from _protection_config import apply_protection_blocks
from _type_patterns_config import build_type_patterns, CN_NUM, cn2int, iter_matches
# ==================== 批量模式 Excel 配置 ====================
# 输入：从哪个 Excel 读取需要分析的 law_id
BATCH_INPUT_FILE = "3168条 定性入库.xlsx"       # 输入 Excel 路径
BATCH_INPUT_SHEET = "Sheet1"                   # 输入 sheet 名
BATCH_INPUT_COLUMN = "law_id"                  # law_id 所在列的表头名称

# 输出：分析结果写入哪个 Excel
BATCH_OUTPUT_FILE = "507测试拆分类型测试.xlsx"     # 输出 Excel 路径（不存在则新建）

# 性能：低配主机调大此值以降低 CPU 占用（秒），0 为不等待
BATCH_DELAY_SECONDS = 0.05                       # 每条分析后的等待秒数
# =============================================================

# ==================== 保护块机制（避免日期、联系方式被误识别为拆分类型） ====================

def is_plain_text(results):
    """全部拆分类型 count==0 或全部 group_count==0 时视为纯文本。"""
    all_zero_count = all(entry["count"] == 0 for entry in results.values())
    no_groups = all(entry["group_count"] == 0 for entry in results.values())
    return all_zero_count or no_groups


def count_paragraphs(text):
    """按换行符分隔，统计非空段落数。"""
    return len([p for p in text.split('\n') if p.strip()])


def count_effective_paragraphs(text):
    """先应用保护块，再统计有效段落数。
    仅包含保护块占位符的行不计入段落数。"""
    protected, _ = apply_protection_blocks(text)
    lines = protected.split('\n')
    effective = 0
    for line in lines:
        stripped = line.strip()
        if not stripped:
            continue
        if re.match(r'^___PB_\w+_\d+___$', stripped):
            continue
        effective += 1
    return max(effective, 1)  # 至少为 1 防止除零



# ---- 全部拆分类型 ----
ALL_SPLIT_TYPES = [
    "条", "章", "节", "编", "部分",
    "括号", "中文顿号", "数字顿号", "数字空格",
    "数字条", "数字章", "数字节",
    "数字点", "数字点点", "数字直连中文",
    "中文是", "要素数字冒号",
]

# 模块级预编译 — 所有类型 pattern 只编译一次，analyze 直接复用
_ALL_PATTERNS = build_type_patterns(ALL_SPLIT_TYPES)


def find_logical_groups(ordinals):
    """
    从一组序数中找出从 1 开始的最大连续逻辑递增组。
    仅用去重后的序数判断「最大能走多远」（max N）。
    返回 max_n（从 1 开始的最大连续递增长度）。
    无有效组时返回 0。
    仅 N >= 2 时视为有效组。
    """
    if not ordinals:
        return 0
    sorted_ords = sorted(set(ordinals))
    best_n = 0
    best_seq = []
    i = 0
    while i < len(sorted_ords):
        if sorted_ords[i] == 1:
            j = i + 1
            expected = 2
            seq = [1]
            while j < len(sorted_ords):
                if sorted_ords[j] == expected:
                    seq.append(expected)
                    expected += 1
                    j += 1
                elif sorted_ords[j] < expected:
                    j += 1
                else:
                    break
            n = expected - 1
            if n >= 2 and n > best_n:
                best_n = n
                best_seq = seq
            i = j
        else:
            i += 1
    return best_n


def count_group_restarts(ordinals_in_order):
    """
    在原始顺序的序数中，统计独立的 1→N 组重启次数。
    每次 ordinal=1 重启且其后存在至少 2 个连续递增序数（即 N≥3），
    视为一个独立组。

    逻辑：扫描原始序列，遇到 1 时向前收集连续递增序数。
    若收集到 ≥3 个不同序数，记录一组。跳过被当前组 span 覆盖的
    后续位置，继续扫描。

    返回: group_count, groups_detail[(1, N), ...]
    """
    if not ordinals_in_order:
        return 0, []

    n_total = len(ordinals_in_order)
    groups = []
    i = 0
    while i < n_total:
        if ordinals_in_order[i] == 1:
            # 从当前 1 开始收集
            collected = {1}
            expected = 2
            j = i + 1
            while j < n_total:
                o = ordinals_in_order[j]
                if o == expected:
                    collected.add(o)
                    expected += 1
                elif o == 1:
                    # 遇到下一个 1，组边界
                    break
                elif o < expected:
                    # 重复，跳过
                    pass
                # gap: 继续扫描
                j += 1
            n = len(collected)
            if n >= 2:
                groups.append((1, n))
                i = j  # 跳到组边界之后
                continue
        i += 1

    return len(groups), groups


def analyze(text):
    text, _ = apply_protection_blocks(text)  # 屏蔽日期/电话/公文号/联系方式
    patterns = _ALL_PATTERNS
    results = OrderedDict()

    # ---- 全量收集 + 按位置去重（同一位置保留最长 match） ----
    all_raw = []  # [(name, text, start, end, val), ...]
    for name, pat, func, m in iter_matches(patterns, text):
            try:
                val = func(m)
                if val is None:
                    continue
                all_raw.append((name, m.group(), m.start(), m.end(), val))
            except Exception:
                continue

    # 优先级：数字直连中文 > 数字点 > 其他（按重叠区间去重）
    TYPE_PRIORITY = {"数字直连中文": 0, "数字点": 1, "数字点点": 1}
    def _priority(name):
        return TYPE_PRIORITY.get(name, 2)

    all_raw.sort(key=lambda x: (x[2], _priority(x[0]), -len(x[1])))
    deduped = []
    occupied = []  # [(start, end), ...] 已占用的区间
    for item in all_raw:
        name, g, s, e, v = item
        # 检查是否与已保留的区间重叠
        overlap = False
        for os, oe in occupied:
            if s < oe and e > os:  # 有重叠
                overlap = True
                break
        if not overlap:
            occupied.append((s, e))
            deduped.append(item)

    # 按类型分组构建 entry
    by_type = OrderedDict()
    for name, g, s, e, v in deduped:
        by_type.setdefault(name, []).append((g, s, e, v))

    for name, pat, func in patterns:
        matches = by_type.get(name, [])

        entry = {
            "count": len(matches),
            "ordinals": [v for _, _, _, v in matches],
            "positions": [(s, e) for _, s, e, _ in matches],
            "scalar_groups": [],
            "tuple_groups": [],   # [(prefix_tuple, start, end), ...]
            "max_n": 0,
            "group_count": 0,
        }

        ordinals = entry["ordinals"]
        if not ordinals:
            results[name] = entry
            continue

        # 标量序数分析
        scalar_ords = [o for o in ordinals if isinstance(o, int)]
        if scalar_ords:
            # max_n: 去重后最大连续序列长度（脊椎指标）
            best_n = find_logical_groups(scalar_ords)
            # group_count: 原始顺序中独立重启次数（附生指标）
            gc, groups_detail = count_group_restarts(scalar_ords)
            entry["max_n"] = best_n
            entry["group_count"] = gc
            entry["scalar_groups"] = groups_detail

        # 元组序数分析：按前缀分组，每组独立判断
        tuple_ords = [o for o in ordinals if isinstance(o, tuple)]
        if tuple_ords:
            prefix_groups = OrderedDict()
            for o in tuple_ords:
                prefix = o[:-1]
                last = o[-1]
                prefix_groups.setdefault(prefix, []).append(last)

            best_tuple_n = 0
            for prefix, lasts in prefix_groups.items():
                best_n = find_logical_groups(lasts)
                gc, gd = count_group_restarts(lasts)
                if best_n >= 3:
                    for start, end in gd:
                        entry["tuple_groups"].append((prefix, start, end))
                    if best_n > best_tuple_n:
                        best_tuple_n = best_n
                    entry["group_count"] += gc

            entry["tuple_groups"].sort(key=lambda g: g[2], reverse=True)
            if best_tuple_n > entry["max_n"]:
                entry["max_n"] = best_tuple_n

        results[name] = entry

    # ===== 交错抑制机制（暂时停止 — 打分阶段替代） =====
    if False:
        # ---- 交错抑制 ----
        # 若数字直连中文/数字空格的序数大部分已出现在数字点/数字点点的元组前缀中，
        # 说明这些只是带点编号的顶层前缀，不应独立算一种拆分类型。
        prefix_set = set()
        for tname in ["数字点", "数字点点"]:
            for o in results[tname]["ordinals"]:
                if isinstance(o, tuple):
                    prefix_set.add(o[0])

        for check_name in ["数字直连中文", "数字空格"]:
            entry = results.get(check_name)
            if entry and prefix_set:
                scalar_ords = [o for o in entry["ordinals"] if isinstance(o, int)]
                if scalar_ords:
                    overlap = sum(1 for o in scalar_ords if o in prefix_set)
                    if overlap / len(scalar_ords) > 0.5:
                        entry["suppressed"] = True

        # ---- 数字点/数字点点在条内抑制 ----
    # 若数字点/数字点点的大部分匹配落在条/数字条的区间内，
    # 说明它们只是条内的子编号，不应作为独立拆分类型。
    tiao_positions = []
    for tname in ["条", "数字条"]:
        tiao_positions.extend(results[tname].get("positions", []))
    if tiao_positions:
        tiao_positions.sort(key=lambda x: x[0])
        tiao_spans = []
        for i in range(len(tiao_positions) - 1):
            tiao_spans.append((tiao_positions[i][0], tiao_positions[i+1][0]))
        tiao_spans.append((tiao_positions[-1][0], len(text)))

        for check_name in ["数字点", "数字点点"]:
            entry = results.get(check_name)
            if not entry or entry["count"] == 0:
                continue
            positions = entry.get("positions", [])
            if not positions:
                continue
            inside_count = 0
            for ps, pe in positions:
                for ts, te in tiao_spans:
                    if ts <= ps < te:
                        inside_count += 1
                        break
            if inside_count > 0 and inside_count / len(positions) > 0.5:
                # 豁免：数字点点序列长(max_n≥10)且组数少(≤3)，说明是独立序号结构，不受条内抑制
                if check_name == "数字点点" and entry.get("max_n", 0) >= 10 and entry.get("group_count", 0) <= 3:
                    pass
                else:
                    entry["suppressed"] = True
                    entry["suppress_reason"] = "条内抑制"

    # ---- 数字顿号包容抑制 ----
    dunhao_entry = results.get("数字顿号")
    if dunhao_entry and dunhao_entry["count"] > 0:
        dh_positions = dunhao_entry.get("positions", [])
        if dh_positions:
            PARENT_ORDER = ["中文顿号", "条", "括号", "数字点", "数字点点"]
            for parent_name in PARENT_ORDER:
                parent_entry = results.get(parent_name)
                if not parent_entry or parent_entry["count"] < 2:
                    continue
                parent_positions = parent_entry.get("positions", [])
                if not parent_positions:
                    continue
                parent_positions.sort(key=lambda x: x[0])
                parent_spans = []
                for i in range(len(parent_positions) - 1):
                    parent_spans.append((parent_positions[i][0], parent_positions[i+1][0]))
                parent_spans.append((parent_positions[-1][0], len(text)))

                inside_count = 0
                for ps, pe in dh_positions:
                    for ts, te in parent_spans:
                        if ts <= ps < te:
                            inside_count += 1
                            break
                if inside_count > 0 and inside_count / len(dh_positions) > 0.5:
                    dunhao_entry["suppressed"] = True
                    dunhao_entry["suppress_reason"] = f"包容抑制({parent_name})"
                    break

    # ---- 条包裹抑制 ----
    tiao_entry = results.get("条")
    if tiao_entry and tiao_entry["count"] > 0 and tiao_entry["max_n"] <= 4:
        tiao_positions_list = tiao_entry.get("positions", [])
        if tiao_positions_list:
            PARENT_ORDER_TIAO = ["中文顿号", "括号", "数字点", "数字点点"]
            for parent_name in PARENT_ORDER_TIAO:
                parent_entry = results.get(parent_name)
                if not parent_entry or parent_entry["count"] < 2:
                    continue
                parent_positions = parent_entry.get("positions", [])
                if not parent_positions:
                    continue
                parent_positions.sort(key=lambda x: x[0])
                parent_spans = []
                for i in range(len(parent_positions) - 1):
                    parent_spans.append((parent_positions[i][0], parent_positions[i+1][0]))
                parent_spans.append((parent_positions[-1][0], len(text)))

                inside_count = 0
                for ps, pe in tiao_positions_list:
                    for ts, te in parent_spans:
                        if ts <= ps < te:
                            inside_count += 1
                            break
                if inside_count > 0 and inside_count / len(tiao_positions_list) > 0.5:
                    tiao_entry["suppressed"] = True
                    tiao_entry["suppress_reason"] = f"条包裹抑制({parent_name})"
                    break

    return results


# ====================================================================
#  差分权重法：从拆分类型序列推断各类型的层级索引
# ====================================================================

def infer_type_levels(type_sequence):
    """差分权重法：从拆分后的片段类型序列推断各类型的索引级别。

    核心机制：
      1. 加权包裹 — 统计相邻 A 之间 B 的出现总次数（非去重），
         若 count(A之间B) > count(B之间A) 则 A 是 B 的父级。
         单次出现的类型视为包裹其后所有内容（顶层容器）。
      2. DAG 层级 — 包裹关系建图 → 拓扑排序 → 最长路径深度。
         无边节点回退到平均跨度排序。

    输入: [(seq, split_type), ...] 或 ["章", "条", ...]
    输出: {类型名: 层级}, 0=最外层, 连续整数
    """
    import statistics
    from collections import defaultdict

    # ---- 输入规范化 ----
    if not type_sequence:
        return {}

    if isinstance(type_sequence[0], dict):
        types = [f.get("split_type") for f in type_sequence if f.get("split_type")]
    elif isinstance(type_sequence[0], (list, tuple)):
        types = [t for _, t in type_sequence if t]
    else:
        types = [t for t in type_sequence if t]

    if not types:
        return {}

    unique_types = list(dict.fromkeys(types))
    if len(unique_types) == 1:
        return {unique_types[0]: 0}

    n = len(types)
    all_types = unique_types

    # ---- 1. 转移矩阵 + 位置索引 ----
    T = defaultdict(lambda: defaultdict(int))
    positions = defaultdict(list)
    for i, t in enumerate(types):
        positions[t].append(i)
        if i < n - 1:
            T[t][types[i+1]] += 1

    # ---- 2. 加权包裹矩阵：C[a][b] = b 在相邻 a 对之间出现的总次数 ----
    C = defaultdict(lambda: defaultdict(int))
    for type_a, pos_list in positions.items():
        # 常规：连续出现对之间的包裹
        for i in range(len(pos_list) - 1):
            start, end = pos_list[i], pos_list[i+1]
            for pos_b in range(start + 1, end):
                type_b = types[pos_b]
                if type_b != type_a:
                    C[type_a][type_b] += 1
        # 仅单次出现的类型：其后所有类型视为被其包裹（顶层容器）
        if len(pos_list) == 1:
            start = pos_list[0]
            for pos_b in range(start + 1, n):
                type_b = types[pos_b]
                if type_b != type_a:
                    C[type_a][type_b] += 1

    # ---- 3. 平均跨度（无边节点时的回退信号） ----
    avg_span = {}
    for t in all_types:
        pl = positions[t]
        if len(pl) >= 2:
            gaps = [pl[i+1] - pl[i] for i in range(len(pl) - 1)]
            avg_span[t] = statistics.mean(gaps)
        else:
            avg_span[t] = float(n)

    # ---- 4. 构建 DAG：加权包裹不对称判定父子关系 ----
    in_degree = defaultdict(int)
    children = defaultdict(set)
    for a in all_types:
        for b in all_types:
            if a == b:
                continue
            if C[a][b] > C[b][a]:
                children[a].add(b)
                in_degree[b] += 1

    for t in all_types:
        if t not in in_degree:
            in_degree[t] = 0

    # ---- 5. 拓扑排序 + 最长路径深度 ----
    queue = [t for t in all_types if in_degree[t] == 0]

    # 无边节点：按平均跨度降序（跨度大=层级高）
    if not queue:
        queue = sorted(all_types, key=lambda t: avg_span.get(t, 0), reverse=True)

    topo = []
    while queue:
        node = queue.pop(0)
        topo.append(node)
        for child in children.get(node, set()):
            in_degree[child] -= 1
            if in_degree[child] == 0:
                queue.append(child)

    for t in all_types:
        if t not in topo:
            topo.append(t)

    levels = {}
    for t in topo:
        max_parent = -1
        for a in all_types:
            if t in children.get(a, set()) and a in levels:
                max_parent = max(max_parent, levels[a])
        levels[t] = max_parent + 1

    # ---- 6. 规范化为连续整数 0..k-1 ----
    unique_levels = sorted(set(levels.values()))
    level_map = {old: new for new, old in enumerate(unique_levels)}
    normalized = {t: level_map[lv] for t, lv in levels.items()}

    return normalized


def _simulate_type_sequence(text, qualifying_types):
    """用指定类型的 pattern 扫描全文，返回按原文位置排序的类型序列。"""
    patterns = build_type_patterns(qualifying_types)
    pairs = []  # [(start_pos, type_name), ...]
    for name, pat, _func in patterns:
        for m in pat.finditer(text):
            pairs.append((m.start(), name))
    pairs.sort(key=lambda x: x[0])
    return [name for _, name in pairs]


def _format_level_chain(levels):
    """将 {类型: 层级} 格式化为可读链。

    同级类型用 | 连接，层级间用 → 连接。
    例: {编:0, 章:1, 节:2, 条:3} → "编→章→节→条" (层级数=4)
    例: {括号:0, 中文顿号:0} → "括号|中文顿号" (层级数=1, 交错)
    返回 (层级链, 层级数)
    """
    if not levels:
        return "-", 0
    from collections import defaultdict
    by_level = defaultdict(list)
    for t, lv in levels.items():
        by_level[lv].append(t)
    max_lv = max(by_level.keys())
    parts = []
    for lv in range(max_lv + 1):
        types_at_lv = sorted(by_level.get(lv, []))
        if types_at_lv:
            parts.append("|".join(types_at_lv))
    return "→".join(parts), max_lv + 1


def format_group_label(name, entry):
    """生成逻辑递增组的可读标签列表"""
    labels = []
    for start, end in entry["scalar_groups"]:
        labels.append(f"1-{end}")
    for prefix, start, end in entry["tuple_groups"]:
        pfx_str = '.'.join(str(x) for x in prefix)
        if pfx_str:
            labels.append(f"{pfx_str}.{start}-{end}")
        else:
            labels.append(f"{start}-{end}")
    return labels


def print_report(results, text="", law_id=None, quiet=False):
    _p = (lambda *a, **kw: None) if quiet else print

    _p("=" * 72)
    _p("  原文拆分类型分析 & 逻辑递增组判断")
    if law_id:
        _p(f"  law_id: {law_id}")
    _p("=" * 72)

    if text:
        preview = text[:250].replace('\n', '\\n')
        # 清洗无法被终端编码的字符
        preview = preview.encode(sys.stdout.encoding or 'utf-8', errors='replace').decode(sys.stdout.encoding or 'utf-8')
        if len(text) > 250:
            preview += "..."
        _p(f"\n  原文 ({len(text)} 字符): {preview}\n")

    # ---- 纯文本处理 ----
    if is_plain_text(results):
        para_count = count_paragraphs(text) if text else 0
        char_count = len(text) if text else 0
        _p(f"  >>> 纯文本 — 未匹配到任何拆分类型\n")
        _p(f"  字符数: {char_count}")
        _p(f"  段落数: {para_count}")

        if char_count < 1000:
            _p(f"  >>> 字数 < 1000，不拆分")
            return {
                "spine_types": ["纯文本"],
                "satellite_types": [],
                "all_tags": ["纯文本"],
                "max_n": 0, "max_gc": 0,
                "is_plain": True,
                "char_count": char_count, "para_count": para_count,
            }

        eff_paras = count_effective_paragraphs(text)
        avg_chars = char_count / eff_paras
        _p(f"  有效段落数(保护后): {eff_paras}")
        _p(f"  平均每段: {avg_chars:.0f} 字")

        if avg_chars > 70:
            _p(f"  >>> 平均段落 > 70 字，推荐按段落拆分")
            return {
                "spine_types": ["纯文本段落拆分"],
                "satellite_types": [],
                "all_tags": ["纯文本段落拆分"],
                "max_n": 0, "max_gc": 0,
                "is_plain": True,
                "char_count": char_count, "para_count": para_count,
            }
        else:
            _p(f"  >>> 平均段落 <= 300 字，不拆分")
            return {
                "spine_types": ["纯文本"],
                "satellite_types": [],
                "all_tags": ["纯文本"],
                "max_n": 0, "max_gc": 0,
                "is_plain": True,
                "char_count": char_count, "para_count": para_count,
            }

    # ---- 计数表 ----
    header = f"  {'拆分类型':<14} {'匹配数':<7} {'组数':<5} {'max N':<6} {'序数集 (前8)':<30}"
    _p("-" * 72)
    _p(header)
    _p("-" * 72)

    qualifying = OrderedDict()

    for name in ALL_SPLIT_TYPES:
        entry = results[name]
        cnt = entry["count"]

        if cnt == 0:
            _p(f"  {name:<14} {0:<7} {'-':<5} {'-':<6} {'-':<30}")
            continue

        # 序数集
        ords = entry["ordinals"]
        ord_strs = []
        for o in ords[:8]:
            if isinstance(o, tuple):
                ord_strs.append('.'.join(str(x) for x in o))
            else:
                ord_strs.append(str(o))
        ord_summary = ', '.join(ord_strs)
        if len(ords) > 8:
            ord_summary += f" (+{len(ords) - 8})"

        gc = entry["group_count"]
        mn = entry["max_n"]
        suppressed = entry.get("suppressed", False)

        if not suppressed and cnt >= 3 and (gc >= 2 or (gc >= 1 and mn >= 3)):
            qualifying[name] = {"max_n": mn, "group_count": gc}

        suffix = "  [已抑制]" if suppressed else ""
        _p(f"  {name:<14} {cnt:<7} {str(gc):<5} {str(mn):<6} {ord_summary:<30}{suffix}")

    # 中文顿号特殊规则：max_n >= 2 即可推荐（放宽 gc/max_n 门槛）
    if "中文顿号" not in qualifying:
        e = results["中文顿号"]
        if (not e.get("suppressed", False)
            and e["max_n"] >= 2):
            qualifying["中文顿号"] = {"max_n": e["max_n"], "group_count": e["group_count"]}

    # ---- 定性 ----
    _p("\n" + "=" * 72)
    _p("  定性判断")
    _p("=" * 72)

    if not qualifying:
        _p("\n  未发现任何符合条件的逻辑递增组 (N>=3)。")
        return {
            "spine_types": ["纯文本"],
            "satellite_types": [],
            "all_tags": ["纯文本"],
            "max_n": 0, "max_gc": 0,
            "is_plain": True,
            "char_count": len(text) if text else 0,
            "para_count": count_paragraphs(text) if text else 0,
        }

    max_n = max(q["max_n"] for q in qualifying.values())
    max_gc = max(q["group_count"] for q in qualifying.values())

    spine_types = [n for n, q in qualifying.items() if q["max_n"] == max_n]
    satellite_types = [n for n, q in qualifying.items()
                       if q["group_count"] == max_gc and n not in spine_types]
    close_n = [n for n, q in qualifying.items()
               if q["max_n"] >= max_n - 1 and n not in spine_types]
    close_gc = [n for n, q in qualifying.items()
                if q["group_count"] >= max_gc - 1 and n not in spine_types and n not in satellite_types]

    _p(f"\n  脊椎类型 (序列最长, max N={max_n}): {', '.join(spine_types)}")
    if satellite_types:
        _p(f"  附生类型 (重启最多, {max_gc} 组): {', '.join(satellite_types)}")
    if close_n:
        _p(f"  N 接近类型 (N>={max_n - 1}): {', '.join(close_n)}")
    if close_gc:
        _p(f"  组数接近类型 (>={max_gc - 1} 组): {', '.join(close_gc)}")

    all_tags = sorted(set(qualifying.keys()))

    # ---- 层级依赖扩展 ----
    # 条+章同时存在时，自动拉上节和编（经典四层结构：编→章→节→条）
    expanded = []
    if "条" in all_tags and "章" in all_tags:
        for t in ["节", "编"]:
            if t not in all_tags and results[t]["count"] > 0:
                expanded.append(t)
    if expanded:
        all_tags = sorted(set(all_tags) | set(expanded))

    # ---- 括号碎片化抑制 ----
    suppressed_override = []

    # 规则A：括号 vs 中文顿号 — 括号组数远多于中文顿号(>=3x)且每组很短(<=4) → 括号只是子列表
    if "括号" in all_tags and "中文顿号" in all_tags:
        kuo_gc = results["括号"]["group_count"]
        kuo_mn = results["括号"]["max_n"]
        dun_gc = results["中文顿号"]["group_count"]
        if dun_gc > 0 and kuo_gc >= dun_gc * 3 and kuo_mn <= 4:
            all_tags.remove("括号")
            suppressed_override.append("括号")

    # 规则B：括号在条/数字条内的碎片化 — 括号组数远多于父类型(>=3x)且每组很短(<=4) → 括号只是条内的局部子列表
    if "括号" in all_tags:
        for parent_tp in ["条", "数字条"]:
            if parent_tp not in all_tags:
                continue
            kuo_gc = results["括号"]["group_count"]
            kuo_mn = results["括号"]["max_n"]
            parent_gc = results[parent_tp]["group_count"]
            if parent_gc > 0 and kuo_gc >= parent_gc * 3 and kuo_mn <= 4:
                all_tags.remove("括号")
                suppressed_override.append("括号")
                break

    # ---- 括号回退推荐位 ----
    if "括号" in suppressed_override:
        dot_suppressed = any(
            results[t].get("suppressed", False) and results[t].get("suppress_reason") == "条内抑制"
            for t in ["数字点", "数字点点"]
        )
        if dot_suppressed:
            all_tags = sorted(set(all_tags) | {"括号"})
            suppressed_override.remove("括号")

    # 规则C：中文顿号为括号父级（括号大部分匹配落在中文顿号区间内）→ 中文顿号优先
    if "括号" in all_tags and "中文顿号" in all_tags:
        dun_positions = results["中文顿号"].get("positions", [])
        kuo_positions = results["括号"].get("positions", [])
        if dun_positions and len(kuo_positions) > 0:
            dun_positions.sort(key=lambda x: x[0])
            kuo_positions.sort(key=lambda x: x[0])
            dun_spans = []
            for i in range(len(dun_positions) - 1):
                dun_spans.append((dun_positions[i][0], dun_positions[i+1][0]))
            dun_spans.append((dun_positions[-1][0], float('inf')))
            inside = 0
            for ps, pe in kuo_positions:
                for ts, te in dun_spans:
                    if ts <= ps < te:
                        inside += 1
                        break
            if inside > 0 and inside / len(kuo_positions) > 0.5:
                all_tags.remove("括号")
                suppressed_override.append("括号")

    # 括号有一组max_n>=8 → 解放抑制
    if "括号" in suppressed_override and results["括号"]["max_n"] >= 8:
        all_tags = sorted(set(all_tags) | {"括号"})
        suppressed_override.remove("括号")

    # ---- 枚举型强制推荐 ----
    # 一是/第一部分/要素数字冒号 等枚举标记，max_n>=3 即推荐（放宽 gc 门槛）
    for t in ["中文是", "部分", "要素数字冒号"]:
        if t not in all_tags and results[t]["count"] >= 3 and results[t]["max_n"] >= 3:
            expanded.append(t)
    if expanded:
        all_tags = sorted(set(all_tags) | set(expanded))

    _p(f"\n  >>> 推荐拆分类型: {all_tags}")
    if spine_types:
        _p(f"  脊椎类型: {', '.join(spine_types)} (max N={max_n})")
    if satellite_types:
        _p(f"  附生类型: {', '.join(satellite_types)} ({max_gc} 组)")
    if expanded:
        _p(f"  联动扩展: {', '.join(expanded)}")
    if suppressed_override:
        kuo_gc = results["括号"]["group_count"]
        kuo_mn = results["括号"]["max_n"]
        _p(f"  括号碎片化抑制: {kuo_gc}组, max N={kuo_mn} — 组数过多且序列短，仅为局部子列表")

    return {
        "spine_types": spine_types,
        "satellite_types": satellite_types,
        "all_tags": all_tags,
        "max_n": max_n,
        "max_gc": max_gc,
        "is_plain": False,
    }


def _ensure_pymysql():
    try:
        import pymysql as _pm
        return _pm
    except ImportError:
        print("pymysql 未安装，请执行 pip install pymysql")
        sys.exit(1)


DB_CONFIG = {
    "host": "192.168.1.109",
    "port": 8001,
    "user": "xoops_root",
    "password": "654321",
    "database": "mtai_serv",
}


def _get_db_connection():
    _pm = _ensure_pymysql()
    return _pm.connect(**DB_CONFIG)


def batch_process():
    """从 Excel 批量读取 law_id，逐一分析，输出到 Excel。"""
    try:
        import openpyxl as _xl
    except ImportError:
        print("openpyxl 未安装，请执行 pip install openpyxl")
        return

    # ---- 读取输入 Excel ----
    if not os.path.exists(BATCH_INPUT_FILE):
        print(f"输入文件不存在: {BATCH_INPUT_FILE}")
        return

    wb_in = _xl.load_workbook(BATCH_INPUT_FILE, read_only=True)
    ws_in = wb_in[BATCH_INPUT_SHEET]

    # 找到 law_id 列
    headers = [str(c.value).strip() if c.value else "" for c in ws_in[1]]
    try:
        law_id_col = headers.index(BATCH_INPUT_COLUMN)
    except ValueError:
        print(f"未找到列 '{BATCH_INPUT_COLUMN}'，可用表头: {headers}")
        wb_in.close()
        return

    # 读取所有 law_id
    law_ids = []
    for row in ws_in.iter_rows(min_row=2, max_row=ws_in.max_row, values_only=True):
        lid = str(row[law_id_col]).strip() if row[law_id_col] else ""
        if lid:
            law_ids.append(lid)
    wb_in.close()
    print(f"共读取 {len(law_ids)} 个 law_id，开始批量分析...\n")

    # ---- 逐条分析 ----
    import time
    conn = _get_db_connection()
    all_results = []
    total = len(law_ids)
    t_start = time.time()

    for i, lid in enumerate(law_ids, start=1):
        # 进度行
        pct = i / total * 100
        elapsed = time.time() - t_start
        eta = (elapsed / i) * (total - i) if i > 0 else 0
        bar_width = 20
        filled = int(bar_width * i / total)
        bar = "[" + "=" * filled + " " * (bar_width - filled) + "]"
        status_line = f"\r  {bar} {i}/{total} ({pct:.1f}%)  ETA {eta:.0f}s  {lid[:20]}..."
        print(status_line, end="", flush=True)

        try:
            with conn.cursor() as cursor:
                cursor.execute(
                    "SELECT attachment_url FROM mt_kb_law_metadata WHERE law_id = %s",
                    (lid,))
                res = cursor.fetchone()
                if not res or not res[0]:
                    continue
                text = res[0]
        except Exception:
            continue

        try:
            raw_results = analyze(text)
            qualifying = OrderedDict()
            for name in ALL_SPLIT_TYPES:
                entry = raw_results[name]
                if (not entry.get("suppressed", False)
                    and entry["count"] >= 3
                    and (entry["group_count"] >= 2 or (entry["group_count"] >= 1 and entry["max_n"] >= 3))):
                    qualifying[name] = {"max_n": entry["max_n"], "group_count": entry["group_count"]}

            # 中文顿号特殊规则：max_n >= 2 即可推荐（放宽 gc/max_n 门槛）
            if "中文顿号" not in qualifying:
                e = raw_results["中文顿号"]
                if (not e.get("suppressed", False)
                    and e["max_n"] >= 2):
                    qualifying["中文顿号"] = {"max_n": e["max_n"], "group_count": e["group_count"]}

            if qualifying:
                max_n = max(q["max_n"] for q in qualifying.values())
                max_gc = max(q["group_count"] for q in qualifying.values())
                spine = [n for n, q in qualifying.items() if q["max_n"] == max_n]
                satellite = [n for n, q in qualifying.items()
                              if q["group_count"] == max_gc and n not in spine]
                all_results.append((lid, text, raw_results, qualifying, spine, satellite, max_n, max_gc))
            else:
                char_count = len(text)
                if char_count < 1000:
                    spine_val = ["纯文本"]
                else:
                    eff_paras = count_effective_paragraphs(text)
                    if char_count / eff_paras > 300:
                        spine_val = ["纯文本段落拆分"]
                    else:
                        spine_val = ["纯文本"]
                all_results.append((lid, text, raw_results, {}, spine_val, [], 0, 0))
        except Exception:
            pass

        if BATCH_DELAY_SECONDS > 0:
            time.sleep(BATCH_DELAY_SECONDS)

    conn.close()
    print()  # 换行
    elapsed_total = time.time() - t_start
    print(f"分析完成，耗时 {elapsed_total:.0f}s，有效结果 {len(all_results)} 条。\n")

    # ---- 写入输出 Excel ----
    print(f"写入 {BATCH_OUTPUT_FILE} ...")

    MAX_LEVELS = 6          # 预分配 0-5 级（编/章/节/条/款/项）
    wb_out = _xl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "分析结果"

    # 动态表头：固定列 + 层级列 + 汇总列 + 尾列
    fixed_headers = ["law_id", "文本长度", "脊椎类型", "脊椎maxN", "附生类型", "附生组数", "全部标签", "索引类型"]
    level_headers = [f"{i}级" for i in range(MAX_LEVELS)]
    summary_headers = ["层级数", "层级链"]
    trailing_headers = ["字符数", "段落数"]
    all_headers = fixed_headers + level_headers + summary_headers + trailing_headers
    FIXED_COLS = len(fixed_headers)                          # 1-8
    CHAIN_START = FIXED_COLS + MAX_LEVELS + 1                # 层级数 col=15
    TRAIL_START = CHAIN_START + len(summary_headers)          # 字符数 col=17

    for col, h in enumerate(all_headers, start=1):
        ws_out.cell(row=1, column=col, value=h)

    for ri, (lid, text, raw, qualifying, spine, satellite, max_n, max_gc) in enumerate(all_results, start=2):
        ws_out.cell(row=ri, column=1, value=lid)
        ws_out.cell(row=ri, column=2, value=len(text))
        ws_out.cell(row=ri, column=3, value=", ".join(spine))
        ws_out.cell(row=ri, column=4, value=max_n)
        ws_out.cell(row=ri, column=5, value=", ".join(satellite) if satellite else "")
        ws_out.cell(row=ri, column=6, value=max_gc)
        tags = sorted(set(qualifying.keys()))
        # 层级依赖扩展：条+章 → 拉上节和编
        if "条" in tags and "章" in tags:
            for t in ["节", "编"]:
                if t not in tags and raw[t]["count"] > 0:
                    tags.append(t)
            tags.sort()
        # 括号碎片化抑制
        kuo_suppressed = False
        if "括号" in tags and "中文顿号" in tags:
            kuo_gc = raw["括号"]["group_count"]
            kuo_mn = raw["括号"]["max_n"]
            dun_gc = raw["中文顿号"]["group_count"]
            if dun_gc > 0 and kuo_gc >= dun_gc * 3 and kuo_mn <= 4:
                tags.remove("括号")
                kuo_suppressed = True
        # 括号在条/数字条内的碎片化
        if "括号" in tags:
            for parent_tp in ["条", "数字条"]:
                if parent_tp not in tags:
                    continue
                kuo_gc = raw["括号"]["group_count"]
                kuo_mn = raw["括号"]["max_n"]
                parent_gc = raw[parent_tp]["group_count"]
                if parent_gc > 0 and kuo_gc >= parent_gc * 3 and kuo_mn <= 4:
                    tags.remove("括号")
                    kuo_suppressed = True
                    break
        # 括号回退推荐位
        if kuo_suppressed:
            dot_suppressed = any(
                raw[t].get("suppressed", False) and raw[t].get("suppress_reason") == "条内抑制"
                for t in ["数字点", "数字点点"]
            )
            if dot_suppressed:
                tags.append("括号")
                tags.sort()
        # 中文顿号为括号父级（括号大部分落在中文顿号区间内）→ 中文顿号优先
        if "括号" in tags and "中文顿号" in tags and not kuo_suppressed:
            dun_pos = raw["中文顿号"].get("positions", [])
            kuo_pos = raw["括号"].get("positions", [])
            if dun_pos and kuo_pos:
                dun_pos.sort(key=lambda x: x[0])
                kuo_pos.sort(key=lambda x: x[0])
                dun_sp = []
                for i in range(len(dun_pos) - 1):
                    dun_sp.append((dun_pos[i][0], dun_pos[i+1][0]))
                dun_sp.append((dun_pos[-1][0], float('inf')))
                inside = 0
                for ps, pe in kuo_pos:
                    for ts, te in dun_sp:
                        if ts <= ps < te:
                            inside += 1
                            break
                if inside > 0 and inside / len(kuo_pos) > 0.5:
                    tags.remove("括号")
                    kuo_suppressed = True

        # 括号有一组max_n>=8 → 解放抑制
        if kuo_suppressed and raw["括号"]["max_n"] >= 6:
            tags.append("括号")
            tags.sort()
            kuo_suppressed = False

        # 枚举型强制推荐：max_n>=3 即推荐
        for t in ["中文是", "部分", "要素数字冒号"]:
            if t not in tags and raw[t]["count"] >= 3 and raw[t]["max_n"] >= 3:
                tags.append(t)
        tags.sort()
        ws_out.cell(row=ri, column=7, value=", ".join(tags))

        # 差分权重法：推断类型索引级别
        if tags and spine not in (["纯文本"], ["纯文本段落拆分"]):
            protected, _ = apply_protection_blocks(text)
            seq = _simulate_type_sequence(protected, set(tags))
            levels = infer_type_levels(seq)
            if levels:
                # col 8: 索引类型（参与分级的全部类型）
                ws_out.cell(row=ri, column=8, value=", ".join(sorted(levels.keys())))
                # col 9-14: 按级别分列
                for lv_num in range(MAX_LEVELS):
                    types_at_lv = sorted(t for t, lv in levels.items() if lv == lv_num)
                    if types_at_lv:
                        ws_out.cell(row=ri, column=FIXED_COLS + 1 + lv_num,
                                    value=", ".join(types_at_lv))
                # col 15: 层级数, col 16: 层级链
                chain_str, level_count = _format_level_chain(levels)
                ws_out.cell(row=ri, column=CHAIN_START, value=level_count)
                ws_out.cell(row=ri, column=CHAIN_START + 1, value=chain_str)

        if spine == ["纯文本"] or spine == ["纯文本段落拆分"]:
            ws_out.cell(row=ri, column=TRAIL_START, value=len(text))
            ws_out.cell(row=ri, column=TRAIL_START + 1, value=count_paragraphs(text))

    wb_out.save(BATCH_OUTPUT_FILE)
    wb_out.close()
    print(f"结果已保存至 {BATCH_OUTPUT_FILE} ({len(all_results)} 条)")


if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "--batch":
        batch_process()
        sys.exit(0)

    law_id = None
    text = None

    if len(sys.argv) > 1:
        if sys.argv[1] == "--id" and len(sys.argv) > 2:
            # 从数据库按 law_id 拉取
            law_id = sys.argv[2]
            conn = _get_db_connection()
            try:
                with conn.cursor() as cursor:
                    cursor.execute(
                        "SELECT attachment_url FROM mt_kb_law_metadata WHERE law_id = %s",
                        (law_id,))
                    res = cursor.fetchone()
                    if res and res[0]:
                        text = res[0]
                    else:
                        print(f"未找到 law_id: {law_id}")
                        sys.exit(1)
            finally:
                conn.close()
        else:
            with open(sys.argv[1], 'r', encoding='utf-8') as f:
                text = f.read()
    else:
        text = """
第一条 为了规范市场秩序，制定本法。
第二条 市场准入实行负面清单制度。
第三条 国务院市场监督管理部门负责全国市场监督管理工作。
第四条 市场主体应当依法开展经营活动。
第一章 总则
第一条 目的和依据
第二条 适用范围
第三条 基本原则
第四条 定义
第五条 监管体制
第五条 经营者应当遵守商业道德。
第六章 附则
第二十条 本法自2025年1月1日起施行。
附1 实施细则
一、 关于市场准入
二、 关于经营许可
三、 关于监督管理
四、 关于法律责任
附2 名词解释
一、 市场主体的定义
二、 负面清单的含义
5.1 总述
5.2 适用范围
6.1 一般规定
6.1.1 细则一
6.1.2 细则二
6.1.3 细则三
6.2 特殊规定
7 附则
（一） 申请材料
（二） 审批流程
（三） 公示要求
（四） 异议处理
（五） 复议程序
（六） 执行监督
"""
        print("未指定输入文件，使用内置测试文本。\n")

    results = analyze(text)
    print_report(results, text, law_id=law_id)
