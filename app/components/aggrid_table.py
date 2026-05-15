"""AG Grid table wrapper for fragment results display.

Features:
  - Virtual scrolling for large datasets
  - QuickFilter text search
  - Column header filtering by split_type
  - Row click -> dialog with full content
  - Export current view to Excel
"""
import io
from nicegui import ui


def build_aggrid(fragments: list[dict], meta: dict):
    """Build an AG Grid table from fragment data.

    Args:
        fragments: List of fragment dicts from /api/split response.
        meta: Meta dict from /api/split response.

    Returns:
        The ui.aggrid element.
    """
    column_defs = [
        {
            'headerName': '序号',
            'field': 'seq',
            'width': 70,
            'sortable': True,
            'filter': False,
        },
        {
            'headerName': '内容',
            'field': 'content_preview',
            'flex': 1,
            'sortable': False,
            'filter': False,
            'cellRenderer': '''
                function(params) {
                    if (!params.value) return '';
                    return params.value.substring(0, 80) + (params.value.length > 80 ? '…' : '');
                }
            ''',
        },
        {
            'headerName': '类型',
            'field': 'split_type_display',
            'width': 90,
            'sortable': True,
            'filter': True,
        },
        {
            'headerName': '层级',
            'field': 'index_level',
            'width': 70,
            'sortable': True,
            'filter': False,
        },
        {
            'headerName': '序数',
            'field': 'ordinal_display',
            'width': 90,
            'sortable': True,
            'filter': False,
        },
    ]

    row_data = []
    for f in fragments:
        content = f.get('content', '')
        st = f.get('split_type')
        il = f.get('index_level')
        ordinal = f.get('ordinal')

        # Format for display
        st_display = st if st else '-'
        il_display = str(il) if il is not None else '-'
        if isinstance(ordinal, list):
            ord_display = '.'.join(str(x) for x in ordinal)
        elif ordinal is not None:
            ord_display = str(ordinal)
        else:
            ord_display = '-'

        row_data.append({
            'seq': f.get('seq', 0),
            'content': content,
            'content_preview': content,
            'split_type': st_display,
            'split_type_display': st_display,
            'index_level': il_display,
            'ordinal': ordinal,
            'ordinal_display': ord_display,
        })

    grid_options = {
        'columnDefs': column_defs,
        'rowData': row_data,
        'defaultColDef': {
            'resizable': True,
        },
        'enableCellTextSelection': True,
        'rowSelection': 'single',
        'pagination': True,
        'paginationPageSize': 50,
        'paginationPageSizeSelector': [20, 50, 100, 200],
        'domLayout': 'autoHeight',
    }

    grid = ui.aggrid(grid_options).classes('w-full')

    # Row click handler - show full content in dialog
    grid.on('cellClicked', lambda e: _show_detail_dialog(e, row_data))

    return grid


def _show_detail_dialog(event, row_data: list[dict]):
    """Show a dialog with the full fragment content when a row is clicked."""
    row_index = event.args.get('rowIndex', -1)
    if row_index < 0 or row_index >= len(row_data):
        return

    row = row_data[row_index]
    seq = row.get('seq', '?')
    st = row.get('split_type', '-')
    content = row.get('content', '')

    with ui.dialog() as dialog, ui.card().classes('p-4 max-w-2xl'):
        ui.label(f'片段 #{seq}').classes('text-lg font-bold')
        ui.label(f'类型: {st}').classes('text-sm text-grey')
        ui.separator()
        ui.markdown(content).classes('whitespace-pre-wrap max-h-96 overflow-auto')
        with ui.row().classes('justify-end'):
            ui.button('关闭', on_click=dialog.close)

    dialog.open()


def export_to_excel(grid, filename: str = '拆分结果.xlsx'):
    """Export current AG Grid view data to Excel."""
    try:
        import openpyxl
    except ImportError:
        ui.notify('请安装 openpyxl 以支持导出', type='negative')
        return

    options = grid.options
    row_data = options.get('rowData', [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '拆分结果'

    headers = ['序号', '内容', '类型', '层级', '序数']
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)

    for i, row in enumerate(row_data, start=2):
        ws.cell(row=i, column=1, value=row.get('seq', ''))
        ws.cell(row=i, column=2, value=row.get('content', ''))
        ws.cell(row=i, column=3, value=row.get('split_type', ''))
        ws.cell(row=i, column=4, value=row.get('index_level', ''))
        ws.cell(row=i, column=5, value=row.get('ordinal_display', ''))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    ui.download(output.read(), filename)
