"""File upload component with Excel column selector.

Handles .txt, .docx, .xlsx parsing.
Uploaded/parsed text goes into a shared textarea.
"""
import io
try:
    import cchardet as chardet
except ImportError:
    import chardet
from nicegui import ui

MAX_FILE_SIZE = 50 * 1024 * 1024  # 50 MB


def _detect_encoding(raw: bytes) -> str:
    result = chardet.detect(raw)
    encoding = result.get('encoding', 'utf-8') if result else 'utf-8'
    return encoding or 'utf-8'


def _parse_txt(raw: bytes) -> str:
    encoding = _detect_encoding(raw)
    return raw.decode(encoding, errors='replace')


def _parse_docx(raw: bytes) -> str:
    try:
        from docx import Document
    except ImportError:
        raise ImportError('请安装 python-docx: pip install python-docx')
    doc = Document(io.BytesIO(raw))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return '\n'.join(paragraphs)


def _parse_xlsx_headers(raw: bytes) -> list[str]:
    """Return first-row headers from the Excel file."""
    try:
        import openpyxl
    except ImportError:
        raise ImportError('请安装 openpyxl: pip install openpyxl')
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True)
    ws = wb.active
    headers = [str(c.value) if c.value is not None else '' for c in ws[1]]
    wb.close()
    return headers


def _parse_xlsx_column(raw: bytes, col_index: int) -> str:
    """Extract text from a single column in the Excel file."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True)
    ws = wb.active
    cells = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_index + 1,
                            max_col=col_index + 1, values_only=True):
        val = row[0]
        if val is not None and str(val).strip():
            cells.append(str(val).strip())
    wb.close()
    return '\n'.join(cells)


class FileUpload:
    """Upload area + Excel column selector + shared textarea.

    Usage:
        uploader = FileUpload(on_text_changed=my_handler)
        # uploader.textarea is the shared ui.textarea
    """

    def __init__(self, on_text_changed=None):
        self.on_text_changed = on_text_changed

        with ui.column().classes('w-full gap-2'):
            # Upload area
            self.upload = ui.upload(
                label='拖拽文件到此处 或 点击上传',
                on_upload=self._handle_upload,
                auto_upload=True,
                max_file_size=MAX_FILE_SIZE,
            ).classes('w-full').props('accept=.txt,.docx,.xlsx')

            # Shared textarea
            self.textarea = ui.textarea(
                '法条文本',
                placeholder='上传文件或在此粘贴法条原文…',
                on_change=lambda e: self._on_text_change(),
            ).classes('w-full h-64')

    def _on_text_change(self):
        if self.on_text_changed:
            self.on_text_changed(self.textarea.value or '')

    async def _handle_upload(self, e):
        """Process uploaded file, handle Excel column selection."""
        raw = e.content.read()

        if len(raw) > MAX_FILE_SIZE:
            ui.notify(f'文件过大，超过 50MB 限制', type='negative')
            self.upload.reset()
            return

        ext = (e.name or '').lower()
        parsed = None

        try:
            if ext.endswith('.txt'):
                parsed = _parse_txt(raw)

            elif ext.endswith('.docx'):
                parsed = _parse_docx(raw)

            elif ext.endswith('.xlsx'):
                # Detect headers and ask user to pick column
                headers = _parse_xlsx_headers(raw)
                if not headers:
                    ui.notify('Excel 文件第一行为空，无法读取列', type='negative')
                    self.upload.reset()
                    return

                # Show column selection dialog
                col_index = await _show_column_dialog(headers)
                if col_index is None:
                    # User cancelled — keep current text, do nothing
                    self.upload.reset()
                    return
                parsed = _parse_xlsx_column(raw, col_index)

            else:
                ui.notify(f'不支持的格式: {ext}', type='negative')
                self.upload.reset()
                return

        except Exception as ex:
            ui.notify(f'文件解析失败: {ex}', type='negative')
            self.upload.reset()
            return

        if not parsed or not parsed.strip():
            ui.notify('文件内容为空', type='warning')
            self.upload.reset()
            return

        # Confirm overwrite if textarea has existing content
        current = (self.textarea.value or '').strip()
        if current:
            result = await _confirm_overwrite()
            if not result:
                self.upload.reset()
                return

        self.textarea.value = parsed
        self._on_text_change()
        self.upload.reset()


async def _show_column_dialog(headers: list[str]) -> int | None:
    """Show dialog for column selection. Returns column index or None if cancelled."""
    result = {'index': None}

    with ui.dialog() as dialog, ui.card().classes('p-4'):
        ui.label('选择要读取的文本列').classes('text-lg font-bold')
        ui.label(f'检测到 {len(headers)} 列').classes('text-sm text-grey')

        options = {f'{h} (第{i+1}列)': i
                   for i, h in enumerate(headers) if h}
        if not options:
            options = {f'第{i+1}列': i for i in range(len(headers))}

        col_select = ui.select(
            options=options,
            value=list(options.values())[0],
            label='文本列',
        ).classes('w-full')

        with ui.row().classes('gap-2'):
            ui.button('确定', on_click=lambda: _set_and_close(0))
            ui.button('取消', on_click=lambda: _set_and_close(None))

        def _set_and_close(val):
            if val == 0:
                result['index'] = col_select.value
            else:
                result['index'] = None
            dialog.close()

    await dialog
    return result['index']


async def _confirm_overwrite() -> bool:
    """Ask user to confirm text overwrite. Returns True to proceed."""
    result = {'ok': False}

    with ui.dialog() as dialog, ui.card().classes('p-4'):
        ui.label('当前内容将被替换').classes('text-lg font-bold')
        ui.label('文本编辑框中已有内容，是否继续？').classes('text-sm text-grey')
        with ui.row().classes('gap-2'):
            ui.button('继续', on_click=lambda: _ok_and_close(dialog, result))
            ui.button('取消', on_click=lambda: dialog.close())

    def _ok_and_close(d, r):
        r['ok'] = True
        d.close()

    await dialog
    return result['ok']
