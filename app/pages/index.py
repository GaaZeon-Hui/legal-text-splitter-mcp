"""Main page: file upload, text editing, parameters, and split trigger."""
import os
from nicegui import app, ui

from app.components.file_upload import FileUpload, _parse_txt, MAX_FILE_SIZE
from app.components.service_client import client as svc, ServiceError


def build():
    """Build the main page layout."""
    # -- Per-page state --
    current_text = ''
    service_online = False

    # -- Header --
    with ui.header().classes('bg-grey-10 text-white'):
        with ui.row().classes('w-full items-center justify-between px-4 py-2'):
            with ui.row().classes('items-center gap-3'):
                ui.icon('description').classes('text-2xl')
                ui.label('法规文本拆分系统').classes('text-xl font-bold')
                ui.button(icon='refresh', on_click=lambda: ui.navigate.reload()) \
                    .props('flat text-white').tooltip('刷新')
            with ui.row().classes('items-center gap-2'):
                _status_dot = ui.element('span').classes('w-2 h-2 rounded-full')
                _status_label = ui.label('检测中…').classes('text-xs text-grey-4')

    # -- Main content --
    with ui.column().classes('w-full max-w-3xl mx-auto p-4 gap-4'):

        def on_text_change(text):
            nonlocal current_text
            current_text = text
            _split_btn.enabled = bool(text.strip()) and service_online

        current_law_ids = []

        def on_law_ids_changed(law_ids):
            nonlocal current_law_ids
            current_law_ids = law_ids
            _split_btn.enabled = bool(law_ids) and service_online
            _split_btn.props('label=批量拆分')

        uploader = FileUpload(on_text_changed=on_text_change,
                              on_law_ids_changed=on_law_ids_changed)

        # 页面刷新时清空旧状态
        ui.timer(0.05, lambda: (
            setattr(uploader.textarea, 'value', ''),
            uploader.textarea.update(),
        ), once=True)

        # -- Split button --
        with ui.row().classes('items-center gap-4'):
            _split_btn = ui.button(
                '开始拆分',
                icon='rocket_launch',
                on_click=lambda: _do_split(),
            ).props('unelevated color=deep-orange text-weight-bold')
            _split_btn.enabled = False
            ui.label('Ctrl+Enter').classes('text-sm text-grey-6')

            # Keyboard shortcut
            ui.keyboard(on_key=lambda e: _on_key(e))

        # -- Multi-txt batch upload --
        _txt_files = {}  # filename -> text content
        with ui.expansion('批量 TXT 处理', icon='inventory_2', value=True).classes('w-full'):
            with ui.column().classes('w-full gap-2'):
                with ui.row().classes('gap-2 items-center'):
                    ui.button('选择 TXT 文件', icon='file_open', on_click=lambda: _pick_txt_files()) \
                        .props('unelevated color=blue')
                    _txt_list = ui.label('未选择文件').classes('text-xs text-grey-6')
                _txt_btn = ui.button(
                    '批量拆分并导出 Excel',
                    icon='table_rows',
                    on_click=lambda: _do_txt_batch(),
                ).props('unelevated color=teal text-weight-bold')
                _txt_btn.enabled = False

        def _pick_txt_files():
            nonlocal _txt_files
            from tkinter import Tk, filedialog
            root = Tk()
            root.withdraw()
            root.attributes('-topmost', True)
            paths = filedialog.askopenfilenames(
                title='选择 TXT 文件',
                filetypes=[('文本文件', '*.txt'), ('所有文件', '*.*')],
            )
            root.destroy()
            for path in paths:
                name = os.path.basename(path)
                with open(path, 'rb') as f:
                    raw = f.read()
                txt = _parse_txt(raw)
                if txt and txt.strip():
                    _txt_files[name] = txt
            _refresh_txt_ui()
            if paths:
                ui.notify(f'已加载 {len(paths)} 个文件', type='positive')

        def _refresh_txt_ui():
            names = list(_txt_files.keys())
            if names:
                _txt_list.set_text(f'已选择: {", ".join(names[:5])}{" ..." if len(names)>5 else ""} ({len(names)} 个文件)')
            else:
                _txt_list.set_text('未选择文件')
            _txt_btn.enabled = bool(_txt_files)

        async def _do_txt_batch():
            nonlocal _txt_files
            if not _txt_files:
                return
            files = list(_txt_files.items())
            _txt_btn.visible = False
            _progress_box.classes(remove='hidden')
            all_results = []
            for i, (name, text) in enumerate(files):
                lid = name.rsplit('.', 1)[0]
                try:
                    r = await svc.split(text)
                    all_results.append({
                        'law_id': lid,
                        'fragments': r.get('fragments', []),
                        'meta': r.get('meta', {}),
                        'analysis': {'spine_types': r['meta'].get('spine_types', []),
                                     'satellite_types': [],
                                     'all_tags': r['meta'].get('all_tags', []),
                                     'char_count': r['meta'].get('char_count', 0),
                                     'para_count': 0,
                                     'max_n': 0, 'max_gc': 0},
                    })
                except Exception as exc:
                    all_results.append({'law_id': lid, 'fragments': [], 'meta': {},
                                       'analysis': {}, 'error': str(exc)})
                _progress_bar.set_value((i + 1) / len(files))
                _progress_label.set_text(f'处理中… {i + 1} / {len(files)}')
            merged = {'results': all_results}
            from datetime import datetime
            default_name = datetime.now().strftime('%m%d%H%M') + '_batch'
            excel_bytes = _build_batch_excel(merged)
            _save_with_dialog(excel_bytes, f'{default_name}.xlsx')
            ok = sum(1 for r in all_results if not r.get('error'))
            ui.notify(f'完成: {ok}/{len(all_results)}', type='positive' if ok==len(all_results) else 'warning')
            _txt_btn.visible = True
            _progress_box.classes('hidden')
            _txt_files = {}
            _txt_list.set_text('未选择文件')
            _txt_btn.enabled = False

        # -- Loading indicator --
        _spinner = ui.spinner(size='lg').classes('hidden')
        _progress_box = ui.element('div').classes('w-full hidden')
        with _progress_box:
            _progress_bar = ui.linear_progress().classes('w-full')
            _progress_label = ui.label('').classes('text-sm text-grey')

    # -- Health poll --
    async def _check_health():
        nonlocal service_online
        service_online = await svc.health()
        if service_online:
            _status_dot.classes(remove='bg-red')
            _status_dot.classes('bg-green')
            _status_label.set_text('服务已连接')
        else:
            _status_dot.classes(remove='bg-green')
            _status_dot.classes('bg-red')
            _status_label.set_text('服务断开')
        _split_btn.enabled = bool(current_text.strip() or current_law_ids) and service_online

    async def _safe_check_health():
        try:
            await _check_health()
        except RuntimeError:
            pass

    ui.timer(5.0, _safe_check_health)

    # -- Actions --
    def _on_key(e):
        if e.key == 'enter' and e.action == 'keydown' and e.modifiers.get('ctrl'):
            nonlocal current_text
            if current_text.strip() and service_online:
                _do_split()

    async def _do_split():
        nonlocal current_text, current_law_ids
        text = current_text.strip()
        law_ids = current_law_ids

        if not text and not law_ids:
            return

        _split_btn.visible = False

        try:
            if law_ids:
                await _do_batch_split(law_ids)
            else:
                _spinner.classes(remove='hidden')
                result = await svc.split(text)
                app.storage.user['last_result'] = result
                app.storage.user['last_text'] = text
                ui.navigate.to('/results')

        except ServiceError as e:
            ui.notify(str(e), type='negative')
        except Exception as e:
            ui.notify(f'请求失败: {e}', type='negative')
        finally:
            _split_btn.visible = True
            _spinner.classes('hidden')
            _progress_box.classes('hidden')


    async def _do_batch_split(law_ids):
        BATCH_SIZE = 100
        all_results = []
        total = (len(law_ids) + BATCH_SIZE - 1) // BATCH_SIZE

        _progress_box.classes(remove='hidden')

        for batch_num in range(total):
            start = batch_num * BATCH_SIZE
            batch = law_ids[start:start + BATCH_SIZE]

            try:
                batch_result = await svc.split_by_ids(batch)
                all_results.extend(batch_result.get('results', []))
            except Exception as exc:
                for lid in batch:
                    all_results.append({
                        'law_id': lid,
                        'fragments': [],
                        'meta': {},
                        'analysis': {},
                        'error': str(exc),
                    })

            progress = (batch_num + 1) / total
            _progress_bar.set_value(progress)
            _progress_label.set_text(f'处理中… {batch_num + 1} / {total} 批 ({len(all_results)} 条)')

        merged = {'results': all_results}
        from datetime import datetime
        default_name = datetime.now().strftime('%m%d%H%M')
        excel_bytes = _build_batch_excel(merged)
        _save_with_dialog(excel_bytes, f'{default_name}.xlsx')
        ok_count = sum(1 for r in all_results if not r.get('error'))
        fail_count = len(all_results) - ok_count
        msg = f'导出 {ok_count} 条'
        if fail_count:
            msg += f'，{fail_count} 条失败'
        ui.notify(msg, type='positive' if not fail_count else 'warning')


def _save_with_dialog(data: bytes, default_filename: str):
    """弹出原生保存对话框，写入文件。"""
    import os, asyncio
    from tkinter import Tk, filedialog
    root = Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    path = filedialog.asksaveasfilename(
        defaultextension='.xlsx',
        filetypes=[('Excel 文件', '*.xlsx'), ('所有文件', '*.*')],
        initialfile=default_filename,
    )
    root.destroy()
    if path:
        with open(path, 'wb') as f:
            f.write(data)
        ui.notify(f'已保存: {os.path.basename(path)}', type='positive')


def _build_batch_excel(batch_result: dict) -> bytes:
    """Build Excel from batch results, matching engine _write_excel format."""
    import io
    import openpyxl

    wb = openpyxl.Workbook()

    ws_a = wb.active
    ws_a.title = '拆分类型分析'
    a_headers = ['law_id', '文本长度', '脊椎类型', '脊椎maxN',
                 '附生类型', '附生组数', '全部标签', '最终拆分类型',
                 '拆分片段数', '字符数', '段落数', '错误信息']
    for col, h in enumerate(a_headers, start=1):
        ws_a.cell(row=1, column=col, value=h)

    ws_s = wb.create_sheet('拆分结果')
    s_headers = ['组', '序号', '内容', '保留列', '索引级别']
    for col, h in enumerate(s_headers, start=1):
        ws_s.cell(row=1, column=col, value=h)

    row_a = 2
    row_s = 2
    for r in batch_result['results']:
        lid = r['law_id']
        a = r.get('analysis') or {}
        meta = r.get('meta') or {}
        error = r.get('error', '')

        ws_a.cell(row=row_a, column=1, value=lid)
        ws_a.cell(row=row_a, column=2, value=a.get('char_count', 0))
        ws_a.cell(row=row_a, column=3, value=', '.join(a.get('spine_types', [])))
        ws_a.cell(row=row_a, column=4, value=a.get('max_n', 0))
        ws_a.cell(row=row_a, column=5, value=', '.join(a.get('satellite_types', [])))
        ws_a.cell(row=row_a, column=6, value=a.get('max_gc', 0))
        ws_a.cell(row=row_a, column=7, value=', '.join(a.get('all_tags', [])))
        ws_a.cell(row=row_a, column=8, value=', '.join(meta.get('all_tags', [])))
        ws_a.cell(row=row_a, column=9, value=meta.get('fragment_count', 0))
        ws_a.cell(row=row_a, column=10, value=a.get('char_count', 0))
        ws_a.cell(row=row_a, column=11, value=a.get('para_count', 0))
        ws_a.cell(row=row_a, column=12, value=error)
        row_a += 1

        for frag in r.get('fragments', []):
            ws_s.cell(row=row_s, column=1, value=lid)
            ws_s.cell(row=row_s, column=2, value=frag.get('seq', ''))
            ws_s.cell(row=row_s, column=3, value=frag.get('content', ''))
            ws_s.cell(row=row_s, column=4, value=frag.get('extra', ''))
            il = frag.get('index_level')
            ws_s.cell(row=row_s, column=5, value=il if il is not None else '')
            row_s += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()
