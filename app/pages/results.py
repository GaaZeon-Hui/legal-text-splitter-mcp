"""Results page: summary bar + pretext-powered table."""
import json
from nicegui import app, ui


def build():
    result = app.storage.user.get('last_result')
    if not result:
        with ui.column().classes('w-full items-center p-8'):
            ui.label('没有拆分结果，请返回主页重新拆分').classes('text-lg text-grey')
            ui.button('返回主页', on_click=lambda: ui.navigate.to('/'))
        return

    fragments = result.get('fragments', [])
    meta = result.get('meta', {})

    with ui.header().classes('bg-primary text-white'):
        with ui.row().classes('w-full items-center justify-between p-2'):
            with ui.row().classes('items-center gap-2'):
                ui.button(icon='arrow_back', on_click=lambda: ui.navigate.to('/')) \
                    .props('flat text-white')
                ui.label('拆分结果').classes('text-xl font-bold')
            ui.button('导出 Excel', icon='download', on_click=lambda: _do_export()) \
                .props('flat text-white')

    with ui.column().classes('w-full p-4 gap-4'):
        with ui.row().classes('w-full flex-wrap gap-4 items-center bg-grey-1 p-3 rounded-lg'):
            ui.label(f'字符数: {meta.get("char_count", 0):,}').classes('text-sm')
            ui.label(f'片段数: {meta.get("fragment_count", 0):,}').classes('text-sm font-bold')
            ui.label(f'类型: {", ".join(meta.get("all_tags", [])) or "-"}').classes('text-sm')
            ui.label(f'层级: {meta.get("level_chain", "-")}').classes('text-sm')
            ui.label(f'耗时: {meta.get("processing_ms", 0)}ms').classes('text-sm')
            ui.label(f'算法: {meta.get("algorithm", "-")}').classes('text-sm')

        if fragments:
            _build_pretext_table(fragments)
        else:
            ui.label('未能拆分出片段').classes('text-grey')


def _build_pretext_table(fragments):
    # Container div
    ui.html(
        '<link rel="stylesheet"'
        ' href="https://fonts.googleapis.com/css2?family=Inter:400,500,600&display=swap">'
        '<div id="pt-root" class="pt-w"></div>'
    ).classes('w-full')

    # CSS injected via add_head_html — guaranteed to be in DOM before JS renders table
    ui.add_head_html('''
    <style>
    #pt-root .pt-w { width: 100%; overflow-x: auto; }
    #pt-root .pt-t {
        width: 100%; border-collapse: collapse; table-layout: fixed !important;
        font-family: "Inter", -apple-system, BlinkMacSystemFont, "Segoe UI",
                     "PingFang SC", "Microsoft YaHei", "Noto Sans SC", sans-serif;
        font-size: 15px; line-height: 1.55; -webkit-font-smoothing: antialiased;
        color: #1a1a1a;
    }
    #pt-root .pt-t th, #pt-root .pt-t td {
        padding: 8px 12px; vertical-align: middle !important;
        border: 1px solid #ddd !important;
    }
    #pt-root .pt-t thead th {
        font-weight: 500; font-size: 12px; color: #888; background: #fafafa;
    }
    #pt-root .pt-t tbody td {
        overflow: hidden; text-overflow: ellipsis; white-space: nowrap;
    }
    #pt-root .pt-t tbody tr { cursor: pointer; }
    #pt-root .pt-t tbody tr:hover td { background: #f6f6f9; }
    #pt-root .pt-c1 { text-align: center !important; font-weight: 700 !important; white-space: nowrap; }
    #pt-root .pt-c2 { text-align: left !important; }
    #pt-root .pt-c3, #pt-root .pt-c4, #pt-root .pt-c5 { text-align: center !important; white-space: nowrap; }
    /* column width enforcement */
    #pt-root .pt-t th:nth-child(1), #pt-root .pt-t td:nth-child(1) { width: 52px !important; }
    #pt-root .pt-t th:nth-child(3), #pt-root .pt-t td:nth-child(3) { width: 72px !important; }
    #pt-root .pt-t th:nth-child(4), #pt-root .pt-t td:nth-child(4) { width: 56px !important; }
    #pt-root .pt-t th:nth-child(5), #pt-root .pt-t td:nth-child(5) { width: 80px !important; }
    </style>
    ''')

    # Build row data
    rows_js = []
    detail_map = {}
    for f in fragments:
        seq = f.get('seq', '')
        content = f.get('content', '')
        st = f.get('split_type') or '-'
        il = f.get('index_level')
        il_str = str(il) if il is not None else '-'
        ord_val = _fmt_ordinal(f.get('ordinal'))
        rows_js.append([seq, content, st, il_str, ord_val])
        detail_map[str(seq)] = {'s': seq, 't': st, 'c': content}

    rows_json = json.dumps(rows_js, ensure_ascii=False)
    detail_json = json.dumps(detail_map, ensure_ascii=False)

    # Set data on window, then load the renderer module
    ui.timer(0.2, lambda: (
        ui.run_javascript(
            f'window.__ROWS = {rows_json};'
            f'window.__DETAIL = {detail_json};'
        ),
        ui.run_javascript(
            'var s=document.createElement("script");'
            's.type="module";'
            's.src="/static/table-renderer.js";'
            'document.body.appendChild(s);'
        ),
    ), once=True)

    # Poll for row clicks
    async def _check_click():
        try:
            raw = await ui.run_javascript(
                'var x=typeof __pd!=="undefined"?__pd:null;__pd=null;'
                'return x?JSON.stringify(x):null',
                timeout=0.3)
        except Exception:
            raw = None
        if raw:
            data = json.loads(raw)
            with ui.dialog() as dialog, ui.card().classes('p-4 max-w-3xl'):
                ui.label(f'片段 #{data.get("s","?")}').classes('text-lg font-bold')
                ui.label(f'类型: {data.get("t","-")}').classes('text-sm text-grey')
                ui.separator()
                ui.markdown(data.get('c', '')).classes(
                    'whitespace-pre-wrap max-h-96 overflow-auto')
                with ui.row().classes('justify-end'):
                    ui.button('关闭', on_click=dialog.close)
            dialog.open()

    ui.timer(0.3, _check_click)


def _fmt_ordinal(ordinal):
    if isinstance(ordinal, list):
        return '.'.join(str(x) for x in ordinal)
    if ordinal is not None:
        return str(ordinal)
    return '-'


def _do_export():
    result = app.storage.user.get('last_result', {})
    fragments = result.get('fragments', [])
    import io
    try:
        import openpyxl
    except ImportError:
        ui.notify('请安装 openpyxl 以支持导出', type='negative')
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '拆分结果'
    for col, h in enumerate(['序号', '内容', '类型', '层级', '序数'], start=1):
        ws.cell(row=1, column=col, value=h)
    for i, frag in enumerate(fragments, start=2):
        ws.cell(row=i, column=1, value=frag.get('seq', ''))
        ws.cell(row=i, column=2, value=frag.get('content', ''))
        ws.cell(row=i, column=3, value=frag.get('split_type', '-'))
        ws.cell(row=i, column=4, value=frag.get('index_level', '-'))
        ws.cell(row=i, column=5, value=_fmt_ordinal(frag.get('ordinal')))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    ui.download(output.read(), '拆分结果.xlsx')
