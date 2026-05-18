"""
流水线脚本 — 拆分类型识别 + 拆分执行一站式

一次DB下载，两道工序：
  1. analyze() 识别拆分类型 → print_report(quiet) 获取 all_tags
  2. all_tags → SPLIT_TYPES → post-类型拆分 执行拆分+回卷

用法:
  python pipeline_split.py                     内置测试文本
  python pipeline_split.py --id <law_id>        单条law_id（从DB）
  python pipeline_split.py --ids <id1> <id2>…   指定多条law_id，每条输出 <law_id>.xlsx
  python pipeline_split.py --batch              批量（从Excel读取law_id列表）
"""
import sys
import re
import os
import time
import json
import traceback
from collections import OrderedDict
from importlib import util

try:
    import openpyxl
except ImportError:
    pass
try:
    import pymysql
except ImportError:
    pass

# ====================================================================
#  配置区
# ====================================================================
 # ---- OpenBLAS 线程限制（防止 numpy 底层库内存分配失败） ----

os.environ.setdefault("OPENBLAS_NUM_THREADS", "2")
os.environ.setdefault("OMP_NUM_THREADS", "2")

# --- 数据源 ---
# "db"  = 从数据库读取全部 law_id（--all）
# "excel" = 从 Excel 读取 law_id 列表（--batch）
LAW_ID_SOURCE = "excel"

# Excel 输入（仅 LAW_ID_SOURCE = "excel" 时生效）
LAW_ID_EXCEL_FILE = "506.xlsx"
LAW_ID_EXCEL_SHEET = "Sheet1"
LAW_ID_EXCEL_COLUMN = "law_id"

# --- 输出 ---
OUTPUT_EXCEL = "506-result-7.xlsx"
SAVE_JSON_SNAPSHOT = True                   # 每 N 条存 JSON 快照（防崩溃）
SNAPSHOT_INTERVAL = 50                      # 快照间隔（条）
BATCH_DELAY_SECONDS = 0.02                  # 每条处理后的等待秒数

# ====================================================================

# ---- 加载组件 ----
import os as _os

_HERE = _os.path.dirname(_os.path.abspath(__file__)) if '__file__' in dir() else _os.getcwd()

_spec_analyze = util.spec_from_file_location(
    '_analyze', _os.path.join(_HERE, 'analyze_split_types.py'))
_mod_analyze = util.module_from_spec(_spec_analyze)
_spec_analyze.loader.exec_module(_mod_analyze)

_spec_post = util.spec_from_file_location(
    '_post', _os.path.join(_HERE, 'post-类型拆分.py'))
_mod_post = util.module_from_spec(_spec_post)
_spec_post.loader.exec_module(_mod_post)

DB_CONFIG = {
    "host": "192.168.1.109",
    "port": 8001,
    "user": "xoops_root",
    "password": "654321",
    "database": "mtai_serv",
}

analyze = _mod_analyze.analyze
print_report = _mod_analyze.print_report
count_paragraphs = _mod_analyze.count_paragraphs
infer_type_levels = _mod_analyze.infer_type_levels
clean_html = _mod_post.clean_html
_apply_protection = _mod_post._apply_protection_blocks
_restore_placeholders = _mod_post._restore_placeholders
split_single_group_with_rollback = _mod_post.split_single_group_with_rollback
split_plain_by_paragraphs = _mod_post.split_plain_by_paragraphs


# ====================================================================
#  数据库
# ====================================================================

def _ensure_pymysql():
    try:
        import pymysql as _pm
        return _pm
    except ImportError:
        print("pymysql 未安装，请执行 pip install pymysql")
        sys.exit(1)


def _get_db_connection():
    _pm = _ensure_pymysql()
    return _pm.connect(**DB_CONFIG)


def _fetch_text(cursor, law_id):
    cursor.execute(
        "SELECT attachment_url FROM mt_kb_law_metadata WHERE law_id = %s",
        (law_id,))
    res = cursor.fetchone()
    return res[0] if res and res[0] else None


# ====================================================================
#  单条流水线
# ====================================================================

def process_single_law(law_id, conn, quiet=True):
    """一条法律的完整流水线：拉取 → 分析 → 决策 → 拆分。

    返回 dict:
        {law_id, analysis, split_types, split_count, split_results, error}
    """
    result = {
        "law_id": law_id,
        "analysis": None,
        "split_types": [],
        "split_count": 0,
        "split_results": [],
        "error": None,
    }

    try:
        # 1. 从DB拉取原始文本
        with conn.cursor() as cursor:
            raw_text = _fetch_text(cursor, law_id)
        if not raw_text:
            result["error"] = "未获取到文本"
            return result

        # 2. 分析拆分类型 — 走完整的 print_report(quiet=True) 拿到所有后处理规则的结果
        raw_results = analyze(raw_text)
        analysis_report = print_report(raw_results, raw_text, law_id=law_id, quiet=quiet)
        result["analysis"] = analysis_report

        all_tags = analysis_report.get("all_tags", [])
        # 拆分顺序：脊椎类型 → 附生类型 → 其余
        spine = analysis_report.get("spine_types", [])
        satellite = analysis_report.get("satellite_types", [])
        ordered = [t for t in (spine + satellite) if t in all_tags]
        remaining = [t for t in all_tags if t not in ordered]
        all_tags = ordered + remaining
        result["split_types"] = all_tags

        # 3. 清洗 + 保护 + 拆分
        cleaned_text = clean_html(raw_text)
        cleaned_text, prot_blocks = _apply_protection(cleaned_text)

        if "纯文本段落拆分" in all_tags:
            # 纯文本走独立段落拆分
            paragraphs = split_plain_by_paragraphs(cleaned_text)
            gdata = []
            for idx, content in enumerate(paragraphs, start=1):
                gdata.append({
                    "group": law_id,
                    "seq": idx,
                    "content": content,
                    "extra": None,
                    "source_id": 0,
                    "split_type": None,
                })
            other_types = [t for t in all_tags if t != "纯文本段落拆分"]
            if other_types:
                gdata = split_single_group_with_rollback(
                    gdata, law_id, split_types_override=other_types, verbose=False)
        elif all_tags and all_tags != ["纯文本"]:
            # 有拆分类型：创建初始 group_data，传入 split_types_override
            gdata = [{
                "group": law_id,
                "seq": 1,
                "content": cleaned_text,
                "extra": None,
                "source_id": 0,
                "split_type": None,
            }]
            gdata = split_single_group_with_rollback(
                gdata, law_id, split_types_override=all_tags, verbose=False)
        else:
            # 纯文本（不分拆）：整个文本作为一行输出
            gdata = [{
                "group": law_id,
                "seq": 1,
                "content": cleaned_text,
                "extra": None,
                "source_id": 0,
                "split_type": None,
            }]

        # 4. 还原保护块占位符
        for frag in gdata:
            frag["content"] = _restore_placeholders(frag["content"], prot_blocks)

        # 5. 差分权重法：推断类型索引级别并标记每个片段
        type_levels = infer_type_levels(gdata)
        for frag in gdata:
            st = frag.get("split_type")
            frag["index_level"] = type_levels.get(st) if st else None

        result["split_results"] = gdata
        result["split_count"] = len(gdata)

    except Exception as e:
        result["error"] = f"{type(e).__name__}: {e}\n{traceback.format_exc()}"
        import sys as _sys
        print(f"\n  [FAIL] {law_id} — {type(e).__name__}: {e}", file=_sys.stderr)

    return result


# ====================================================================
#  Excel 输出（对齐 post-类型拆分 格式）
# ====================================================================

def _write_excel(all_results, output_path):
    """将流水线结果写入 Excel（双 sheet）。

    Sheet 1: 拆分类型分析（含最终拆分类型列）
    Sheet 2: 拆分结果（对齐 post-类型拆分 的四列格式）
    """
    wb = openpyxl.Workbook()

    # ---- Sheet 1: 拆分类型分析 ----
    ws_a = wb.active
    ws_a.title = "拆分类型分析"
    a_headers = [
        "law_id", "文本长度", "脊椎类型", "脊椎maxN",
        "附生类型", "附生组数", "全部标签", "最终拆分类型",
        "拆分片段数", "字符数", "段落数", "错误信息",
    ]
    for col, h in enumerate(a_headers, start=1):
        ws_a.cell(row=1, column=col, value=h)

    # ---- Sheet 2: 拆分结果（与 post-类型拆分 格式一致） ----
    ws_s = wb.create_sheet("拆分结果")
    s_headers = ["组", "序号", "内容", "保留列", "索引级别"]
    for col, h in enumerate(s_headers, start=1):
        ws_s.cell(row=1, column=col, value=h)

    row_a = 2
    row_s = 2

    for r in all_results:
        lid = r["law_id"]
        a = r["analysis"] or {}
        spine = a.get("spine_types", [])
        satellite = a.get("satellite_types", [])
        all_tags = a.get("all_tags", [])
        split_types = r.get("split_types", [])
        is_plain = a.get("is_plain", False)

        # 分析 sheet
        ws_a.cell(row=row_a, column=1, value=lid)
        ws_a.cell(row=row_a, column=2, value=a.get("char_count", 0))
        ws_a.cell(row=row_a, column=3, value=", ".join(spine))
        ws_a.cell(row=row_a, column=4, value=a.get("max_n", 0))
        ws_a.cell(row=row_a, column=5, value=", ".join(satellite) if satellite else "")
        ws_a.cell(row=row_a, column=6, value=a.get("max_gc", 0))
        ws_a.cell(row=row_a, column=7, value=", ".join(all_tags))
        ws_a.cell(row=row_a, column=8, value=", ".join(split_types))
        ws_a.cell(row=row_a, column=9, value=r["split_count"])
        if is_plain:
            ws_a.cell(row=row_a, column=10, value=a.get("char_count", 0))
            ws_a.cell(row=row_a, column=11, value=a.get("para_count", 0))
        ws_a.cell(row=row_a, column=12, value=r["error"] or "")
        row_a += 1

        # 拆分结果 sheet
        for frag in r["split_results"]:
            ws_s.cell(row=row_s, column=1, value=frag.get("group", lid))
            ws_s.cell(row=row_s, column=2, value=frag.get("seq", ""))
            content = frag.get("content", "")
            ws_s.cell(row=row_s, column=3, value=content)
            ws_s.cell(row=row_s, column=4, value=frag.get("extra", ""))
            il = frag.get("index_level")
            ws_s.cell(row=row_s, column=5, value=il if il is not None else "")
            row_s += 1

    wb.save(output_path)
    wb.close()


# ====================================================================
#  JSON 快照
# ====================================================================

def _save_snapshot(all_results, snapshot_path):
    snapshot = []
    for r in all_results:
        snapshot.append({
            "law_id": r["law_id"],
            "analysis": r["analysis"],
            "split_types": r["split_types"],
            "split_count": r["split_count"],
            "error": r["error"],
        })
    with open(snapshot_path, "w", encoding="utf-8") as f:
        json.dump(snapshot, f, ensure_ascii=False, indent=2)


def _get_all_law_ids(conn):
    """从数据库获取全部 law_id 列表。"""
    with conn.cursor() as cursor:
        cursor.execute("SELECT law_id FROM mt_kb_law_metadata WHERE attachment_url IS NOT NULL AND attachment_url != ''")
        return [row[0] for row in cursor.fetchall()]


def _read_law_ids_from_excel(filepath, sheet, column):
    """从 Excel 读取 law_id 列表。"""
    if not os.path.exists(filepath):
        print(f"输入文件不存在: {filepath}")
        return []
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb[sheet]
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    try:
        col_idx = headers.index(column)
    except ValueError:
        print(f"未找到列 '{column}'，可用表头: {headers}")
        wb.close()
        return []
    law_ids = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        lid = str(row[col_idx]).strip() if row[col_idx] else ""
        if lid:
            law_ids.append(lid)
    wb.close()
    return law_ids


# ====================================================================
#  主流程
# ====================================================================

def pipeline_run(law_ids):
    if not law_ids:
        print("没有需要处理的 law_id。")
        return

    conn = _get_db_connection()
    all_results = []
    total = len(law_ids)
    t_start = time.time()
    snapshot_path = OUTPUT_EXCEL.replace(".xlsx", "_snapshot.json")

    print(f"共 {total} 条，开始流水线处理...\n")

    for i, lid in enumerate(law_ids, start=1):
        pct = i / total * 100
        elapsed = time.time() - t_start
        eta = (elapsed / i) * (total - i) if i > 0 else 0
        bar_width = 20
        filled = int(bar_width * i / total)
        bar = "[" + "=" * filled + " " * (bar_width - filled) + "]"

        status_line = f"\r  {bar} {i}/{total} ({pct:.1f}%)  ETA {eta:.0f}s  {lid[:20]}..."
        print(status_line, end="", flush=True)

        result = process_single_law(lid, conn)
        all_results.append(result)

        if result["error"]:
            err_msg = result["error"].split('\n')[0]  # 第一行是类型+信息
            print(f"\n  [FAIL] {lid} — {err_msg}")
        elif result["analysis"]:
            tags = ", ".join(result["split_types"])
            extra = f"  [{tags}]  -> {result['split_count']} fragments"
            print(f"\r  {bar} {i}/{total} ({pct:.1f}%)  ETA {eta:.0f}s  {lid[:20]}...{extra}")
        else:
            print()

        if SAVE_JSON_SNAPSHOT and i % SNAPSHOT_INTERVAL == 0:
            _save_snapshot(all_results, snapshot_path)

        if BATCH_DELAY_SECONDS > 0:
            time.sleep(BATCH_DELAY_SECONDS)

    conn.close()

    elapsed_total = time.time() - t_start
    succeeded = sum(1 for r in all_results if not r["error"])
    failed = total - succeeded
    print(f"\n流水线完成，耗时 {elapsed_total:.0f}s，成功 {succeeded}/{total} 条。\n")

    if failed > 0:
        print(f"失败 {failed} 条:")
        for r in all_results:
            if r["error"]:
                err_one_line = r["error"].split('\n')[0]
                print(f"  {r['law_id']} — {err_one_line}")
        print()

    if SAVE_JSON_SNAPSHOT:
        _save_snapshot(all_results, snapshot_path)
        print(f"快照已保存至 {snapshot_path}")

    print(f"写入 {OUTPUT_EXCEL} ...")
    _write_excel(all_results, OUTPUT_EXCEL)
    print(f"结果已保存至 {OUTPUT_EXCEL} ({succeeded} 条)")


def _run_single(law_id):
    """单条 DB 模式 — 带终端输出，复用 process_single_law。"""
    conn = _get_db_connection()
    try:
        result = process_single_law(law_id, conn, quiet=False)
        if result["error"]:
            err_first_line = result["error"].split('\n')[0]
            print(f"\n  [FAIL] {law_id} — {err_first_line}")
            return

        all_tags = result["split_types"]
        gdata = result["split_results"]
        print(f"\n  最终拆分类型: {all_tags}")
        print(f"  拆分完成，共 {len(gdata)} 个片段")

        for frag in gdata[:5]:
            content_preview = frag["content"][:80].replace('\n', '\\n')
            print(f"    [{frag['seq']}] {content_preview}...")
        if len(gdata) > 5:
            print(f"    ... 共 {len(gdata)} 个片段")

        _write_excel([result], OUTPUT_EXCEL)
        print(f"\n  结果已保存至 {OUTPUT_EXCEL}")

    except Exception as e:
        print(f"\n  [FAIL] {law_id} — {type(e).__name__}: {e}")
        traceback.print_exc()
    finally:
        conn.close()


def _run_ids(law_ids):
    """对多条 law_id 逐条完整流水线，输出到 <law_id>.xlsx。"""
    conn = _get_db_connection()
    try:
        for law_id in law_ids:
            print(f"\n{'='*60}")
            print(f"  处理: {law_id}")
            print(f"{'='*60}")
            result = process_single_law(law_id, conn, quiet=False)

            out_path = f"{law_id}.xlsx"
            if result["error"]:
                err_first_line = result["error"].split('\n')[0]
                print(f"\n  [FAIL] {law_id} — {err_first_line}")
            else:
                all_tags = result["split_types"]
                gdata = result["split_results"]
                print(f"\n  最终拆分类型: {all_tags}")
                print(f"  拆分完成，共 {len(gdata)} 个片段")
                for frag in gdata[:5]:
                    content_preview = frag["content"][:80].replace('\n', '\\n')
                    print(f"    [{frag['seq']}] {content_preview}...")
                if len(gdata) > 5:
                    print(f"    ... 共 {len(gdata)} 个片段")

            _write_excel([result], out_path)
            print(f"  结果已保存至 {out_path}")
    finally:
        conn.close()


# ====================================================================
#  公共
# ====================================================================

def _run_test():
    """内置测试：用空白文本走完整流程。"""
    test_text = """



"""
    print("使用内置测试文本。\n")
    test_law_id = "test-001"

    results = analyze(test_text)
    analysis_report = print_report(results, test_text, law_id=test_law_id)
    all_tags = analysis_report.get("all_tags", [])

    print(f"\n  最终拆分类型: {all_tags}")
    print(f"  开始拆分...")

    cleaned_text = clean_html(test_text)

    if all_tags and all_tags != ["纯文本"]:
        gdata = [{
            "group": test_law_id, "seq": 1,
            "content": cleaned_text, "extra": None,
            "source_id": 0, "split_type": None,
        }]
        gdata = split_single_group_with_rollback(
            gdata, test_law_id, split_types_override=all_tags)
    else:
        gdata = [{
            "group": test_law_id, "seq": 1,
            "content": cleaned_text, "extra": None,
            "source_id": 0, "split_type": None,
        }]

    print(f"  拆分完成，共 {len(gdata)} 个片段")

    for frag in gdata[:5]:
        content_preview = frag["content"][:80].replace('\n', '\\n')
        print(f"    [{frag['seq']}] {content_preview}...")
    if len(gdata) > 5:
        print(f"    ... 共 {len(gdata)} 个片段")

    all_results = [{
        "law_id": test_law_id,
        "analysis": analysis_report,
        "split_types": all_tags,
        "split_count": len(gdata),
        "split_results": gdata,
        "error": None,
    }]
    _write_excel(all_results, OUTPUT_EXCEL)
    print(f"\n  结果已保存至 {OUTPUT_EXCEL}")


# ====================================================================
#  入口
# ====================================================================

if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "--all":
        conn = _get_db_connection()
        try:
            law_ids = _get_all_law_ids(conn)
        finally:
            conn.close()
        print(f"从数据库读取到 {len(law_ids)} 个 law_id\n")
        pipeline_run(law_ids)

    elif len(sys.argv) > 1 and sys.argv[1] == "--batch":
        law_ids = _read_law_ids_from_excel(
            LAW_ID_EXCEL_FILE, LAW_ID_EXCEL_SHEET, LAW_ID_EXCEL_COLUMN)
        print(f"从 {LAW_ID_EXCEL_FILE} 读取到 {len(law_ids)} 个 law_id\n")
        pipeline_run(law_ids)

    elif len(sys.argv) > 1 and sys.argv[1] == "--id" and len(sys.argv) > 2:
        _run_single(sys.argv[2])

    elif len(sys.argv) > 1 and sys.argv[1] == "--ids" and len(sys.argv) > 2:
        _run_ids(sys.argv[2:])

    elif len(sys.argv) > 1 and sys.argv[1] == "--test":
        _run_test()

    else:
        test_text = """
第一条 为了规范市场秩序，制定本法。
第二条 市场准入实行负面清单制度。
第三条 国务院市场监督管理部门负责全国市场监督管理工作。
第四条 市场主体应当依法开展经营活动。
第一章 总则
第一条 目的和依据
第二条 适用范围
第三条 基本原则
第四条 定义
第五条 监管体制
第五条 经营者应当遵守商业道德。
第六章 附则
第二十条 本法自2025年1月1日起施行。
附1 实施细则
一、 关于市场准入
二、 关于经营许可
三、 关于监督管理
四、 关于法律责任
附2 名词解释
一、 市场主体的定义
二、 负面清单的含义
5.1 总述
5.2 适用范围
6.1 一般规定
6.1.1 细则一
6.1.2 细则二
6.1.3 细则三
6.2 特殊规定
7 附则
（一） 申请材料
（二） 审批流程
（三） 公示要求
（四） 异议处理
（五） 复议程序
（六） 执行监督
"""
        print(f"配置 LAW_ID_SOURCE = \"{LAW_ID_SOURCE}\"")
        if LAW_ID_SOURCE == "excel":
            law_ids = _read_law_ids_from_excel(
                LAW_ID_EXCEL_FILE, LAW_ID_EXCEL_SHEET, LAW_ID_EXCEL_COLUMN)
            print(f"从 {LAW_ID_EXCEL_FILE} 读取到 {len(law_ids)} 个 law_id\n")
            pipeline_run(law_ids)
        else:
            conn = _get_db_connection()
            try:
                law_ids = _get_all_law_ids(conn)
            finally:
                conn.close()
            print(f"从数据库读取到 {len(law_ids)} 个 law_id\n")
            pipeline_run(law_ids)

