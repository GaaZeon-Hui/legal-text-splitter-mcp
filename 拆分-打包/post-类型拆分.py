import re
import openpyxl
import sys

try:
    import pymysql
except ImportError:
    pass

from _protection_config import apply_protection_blocks as _apply_protection_blocks
from _protection_config import _restore_placeholders
from _type_patterns_config import build_type_patterns, iter_matches, PATH_B_TYPES
from analyze_scored import score_ordinals


def split_plain_by_paragraphs(text):
    """保护块应用后按换行符拆分纯文本，纯占位符行不单独成段。"""
    protected, _ = _apply_protection_blocks(text)
    lines = protected.split('\n')
    result = []
    for line in lines:
        stripped = line.strip()
        if not stripped:
            continue
        if re.match(r'^龘?___PB_\w+_\d+___$', stripped):
            continue
        result.append(stripped)
    return result


SPLIT_TYPES = [
    "条", "章", "节", "编", "部分",
    "括号", "中文顿号", "数字顿号", "数字空格",
    "数字条", "数字章", "数字节",
    "数字点", "数字点点", "数字直连中文",
    "中文是", "要素数字冒号",
]

# ---- 打分引擎拦截类型：这些类型不走 Path A/B，只走打分 ----
SCORED_TYPES = {
    "条", "数字条", "数字点", "数字点点",
    "数字直连中文", "数字空格", "数字节",
}

# 数据库路径配置：
DB_PRIMARY_TYPE = "条"
DB_SECONDARY_TYPES = None

# =====================================================


# ---- 构建拆分模式 ----
type_patterns = build_type_patterns(SPLIT_TYPES)

# 确保 DB_PRIMARY_TYPE 的 pattern 始终在 type_patterns 中
_type_names_in_patterns = {p[0] for p in type_patterns}
if DB_PRIMARY_TYPE not in _type_names_in_patterns:
    type_patterns.extend(build_type_patterns([DB_PRIMARY_TYPE]))

_TRANSITION_FALLBACK_PATTERNS = [
    ("数字点",   re.compile(r"(\d+(?:\.\d+)+)(?=\s|[一-鿿]|[（(《〈]|$)"),
     lambda m: tuple(int(x) for x in m.group(1).split('.'))),
    ("数字点点", re.compile(r"(\d+(?:\.\d+)*)\.(?=\s|[一-鿿]|[（(《〈]|$)"),
     lambda m: tuple(int(x) for x in m.group(1).split('.')) if '.' in m.group(1) else int(m.group(1))),
    ("数字空格", re.compile(r"(\d+) +"),
     lambda m: int(m.group(1))),
    ("数字直连中文", re.compile(r"(\d+)(?!\.\d)(?=[一-鿿])"),
     lambda m: int(m.group(1))),
]

if not type_patterns: raise ValueError("请至少选择一种拆分关键词类型！")

def get_ordinal(content):
    for _, pat, func in type_patterns:
        m = pat.match(content)
        if m: return func(m)
    return None

def find_split_point_for_types(content, type_names):
    for name, pat, func, m in iter_matches(type_patterns, content, type_names=type_names):
            start = m.start()
            pre = content[:start]
            if pre.strip() != "":
                return pre, content[start:], name
    return None

# ===== 打分路径：拦截 SCORED_TYPES，替代 Path A/B =====
def _scoring_path_combined(group_data, split_type_order, score_collector=None):
    """对所有 SCORED_TYPES 的行合并打分，返回 {tp: set of uid} 标记。

    合并打分保证跨类型的层级关系（如 '3' 与 '3.1'）不丢失。

    若 score_collector 传入，则写入打分明细：
        score_collector[tp] = {"ords": [...], "scores": [...], "kept_mask": [...]}
    """
    scored_used = [tp for tp in split_type_order if tp in SCORED_TYPES]
    if not scored_used:
        return {}

    all_rows = [r for r in group_data if r.get("split_type") in SCORED_TYPES]
    if len(all_rows) < 2:
        return {}

    all_rows.sort(key=lambda x: x["seq"])

    ord_strs = []
    mapped = []
    for r in all_rows:
        o = get_ordinal(r["content"])
        if o is None:
            continue
        mapped.append(r)
        if isinstance(o, tuple):
            ord_strs.append('.'.join(str(x) for x in o))
        else:
            ord_strs.append(str(o))

    if len(ord_strs) < 2:
        return {}

    scores, kept_mask = score_ordinals(ord_strs)

    # 外部收集器：暴露完整打分明细
    if score_collector is not None:
        score_collector["_ords"] = ord_strs
        score_collector["_scores"] = scores
        score_collector["_kept_mask"] = kept_mask
        score_collector["_mapped"] = mapped

    marks = {}
    for r, keep in zip(mapped, kept_mask):
        if not keep:
            tp = r.get("split_type")
            if tp:
                marks.setdefault(tp, set()).add(r["uid"])
    return marks

# ===== 全局后向回卷 =====
def _ordinal_prefix(ord_val):
    if isinstance(ord_val, tuple):
        return (ord_val[:-1], ord_val[-1])
    return ((), ord_val)


def _prefix_share_root(prefix_a, prefix_b):
    """两个前缀是否有共同首分量（属于同一上层结构）"""
    if not prefix_a or not prefix_b:
        return False
    return prefix_a[0] == prefix_b[0]


def _path_b_mark(tp, group, group_data):
    """路径B：建堆 → 堆延长 → 标记干扰（仅同类型），返回待回卷 uid 集合"""
    # group: [(row, last), ...] 按 seq 已排序

    to_rollback = set()

    # 1. 建堆：按 seq 物理顺序，相邻且序数连续则同堆
    heaps = []
    current = []
    for r, last in group:  # group 已按 seq 排序
        if not current:
            current = [(r, last)]
        elif last == current[-1][1] + 1:
            current.append((r, last))
        else:
            heaps.append(current)
            current = [(r, last)]
    if current:
        heaps.append(current)

    # 标注堆元数据
    heap_info = []
    for h in heaps:
        ordinals = [last for _, last in h]
        seqs = [r["seq"] for r, _ in h]
        heap_info.append({
            "ordinals": ordinals,
            "start_ord": ordinals[0],
            "end_ord": ordinals[-1],
            "min_seq": min(seqs),
            "max_seq": max(seqs),
            "end_seq": seqs[-1],           # end_ord 对应元素的 seq（堆内最后一条）
            "uids": {r["uid"] for r, _ in h},
        })

    # 2. 堆延长：DP 价值驱动的候选选择
    #
    #   只有 start_ord==1 的堆有权延伸。
    #   当存在多个 start_ord == end_ord+1 的候选堆时，对每个候选执行 DP 计算
    #   "best_value" = 自身序数加权和 + 后续链最大序数和，选 best_value 最大的。
    #   这保证选"远期总价值最大"的链，而非简单地按位置近的优先。
    #
    merged_indices = set()
    skip = (tp == "条")

    def _chain_value(h_idx, available, memo):
        """从 heap_info[h_idx] 开始，沿 +1 扩展链能获得的最大序数加权和"""
        key = h_idx
        if key in memo:
            return memo[key]
        h = heap_info[h_idx]
        n = h["end_ord"] - h["start_ord"] + 1
        val = (h["start_ord"] + h["end_ord"]) * n // 2
        nxt = [
            j for j in available if j != h_idx
            and heap_info[j]["start_ord"] == h["end_ord"] + 1
            and heap_info[j]["min_seq"] > h["max_seq"]
        ]
        if nxt:
            val += max(_chain_value(j, available, memo) for j in nxt)
        memo[key] = val
        return val

    for i, h1 in enumerate(heap_info):
        if i in merged_indices:
            continue
        if h1["start_ord"] != 1:
            continue
        while True:
            found = False
            available = {j for j in range(len(heap_info)) if j not in merged_indices}

            # +1 连续延长 — 在所有位置合格的候选中选 best_value 最大的
            cand_plus1 = [
                j for j in available if j != i
                and heap_info[j]["min_seq"] > h1["end_seq"]
                and heap_info[j]["start_ord"] == h1["end_ord"] + 1
            ]
            if cand_plus1:
                memo = {}
                scored = [(j, _chain_value(j, available, memo)) for j in cand_plus1]
                scored.sort(key=lambda x: (-x[1], heap_info[x[0]]["min_seq"]))
                best_j = scored[0][0]
                h2 = heap_info[best_j]
                h1["ordinals"] = sorted(set(h1["ordinals"] + h2["ordinals"]))
                h1["end_ord"] = h2["end_ord"]
                h1["min_seq"] = min(h1["min_seq"], h2["min_seq"])
                h1["max_seq"] = max(h1["max_seq"], h2["max_seq"])
                h1["end_seq"] = h2["end_seq"]
                h1["uids"].update(h2["uids"])
                merged_indices.add(best_j)
                found = True

            if found:
                continue

            # +2 跳过延长 — 仅条、长堆(>=6)、仅一次
            if skip and len(h1["ordinals"]) >= 6:
                cand_plus2 = [
                    j for j in available if j != i
                    and heap_info[j]["min_seq"] > h1["end_seq"]
                    and heap_info[j]["start_ord"] == h1["end_ord"] + 2
                ]
                if cand_plus2:
                    memo = {}
                    scored = [(j, _chain_value(j, available, memo)) for j in cand_plus2]
                    scored.sort(key=lambda x: (-x[1], heap_info[x[0]]["min_seq"]))
                    best_j = scored[0][0]
                    h2 = heap_info[best_j]
                    h1["ordinals"] = sorted(set(h1["ordinals"] + h2["ordinals"]))
                    h1["end_ord"] = h2["end_ord"]
                    h1["min_seq"] = min(h1["min_seq"], h2["min_seq"])
                    h1["max_seq"] = max(h1["max_seq"], h2["max_seq"])
                    h1["end_seq"] = h2["end_seq"]
                    h1["uids"].update(h2["uids"])
                    merged_indices.add(best_j)
                    skip = False
                    found = True

            if not found:
                break

    # 移除被合并的堆
    heap_info = [h for idx, h in enumerate(heap_info) if idx not in merged_indices]

    # 3. 标记干扰元素（仅同类型）
    for heap in heap_info:
        for r in group_data:
            if heap["min_seq"] <= r["seq"] <= heap["max_seq"]:
                if r.get("split_type") == tp and r["uid"] not in heap["uids"]:
                    to_rollback.add(r["uid"])

    # 不在任何堆内的同类型元素也标记
    all_heap_uids = set()
    for heap in heap_info:
        all_heap_uids.update(heap["uids"])
    for r, _ in group:
        if r["uid"] not in all_heap_uids:
            to_rollback.add(r["uid"])

    # 堆首元素不是 ordinal=1 → 整堆标记回卷
    for heap in heap_info:
        if heap["start_ord"] != 1:
            for uid in heap["uids"]:
                to_rollback.add(uid)

    return to_rollback


def global_backward_rollback(group_data, group_name, split_type_order, score_collector=None):
    """
    Phase 1: 每个类型独立建堆+标记（同类型入堆，同类型标记）。
    Phase 2: 按 split_type_order 顺序逐类型回卷（从后往前）。
    """
    group_data.sort(key=lambda x: x["seq"])

    # ---- Phase 1: 每个类型独立建堆 + 标记（同类型入堆，同类型标记） ----
    all_marks = {}  # {tp: set of uids}

    # 打分拦截：SCORED_TYPES 合并打分，跳过 Path A/B
    scored_marks = _scoring_path_combined(group_data, split_type_order, score_collector)
    all_marks.update(scored_marks)

    for tp in split_type_order:
        if tp in SCORED_TYPES:
            continue  # 已由打分引擎处理

        rows = [r for r in group_data if r.get("split_type") == tp]
        if len(rows) < 2:
            continue

        prefix_groups = {}
        all_ords = []
        for r in rows:
            o = get_ordinal(r["content"])
            if o is None:
                continue
            prefix, last = _ordinal_prefix(o)
            all_ords.append(o)
            prefix_groups.setdefault(prefix, []).append((r, last))

        if not prefix_groups:
            continue

        is_path_a = (tp in ("数字点", "数字点点"))
        if tp in ("数字条", "数字章", "数字节"):
            has_tuple = any(isinstance(o, tuple) for o in all_ords)
            is_path_a = has_tuple

        marks = set()

        if is_path_a:
            # 路径A：孤儿判断 + 递增组间隙标记
            for prefix, grp in prefix_groups.items():
                if len(grp) < 2:
                    orphan_row, _ = grp[0]
                    orphan_seq = orphan_row["seq"]
                    prev_row = next((r for r in group_data if r["seq"] == orphan_seq - 1), None)
                    next_row = next((r for r in group_data if r["seq"] == orphan_seq + 1), None)

                    def _get_prefix(row):
                        if row is None: return None
                        o = get_ordinal(row["content"])
                        if o is None: return None
                        return _ordinal_prefix(o)[0]

                    prev_pf = _get_prefix(prev_row)
                    next_pf = _get_prefix(next_row)
                    prev_ok = prev_pf is not None and prev_pf in prefix_groups and len(prefix_groups[prev_pf]) >= 2
                    next_ok = next_pf is not None and next_pf in prefix_groups and len(prefix_groups[next_pf]) >= 2

                    should = False
                    if prev_ok and next_ok:
                        should = True
                    elif not prev_ok and not next_ok:
                        should = True
                    else:
                        nv = prev_pf if not prev_ok else next_pf
                        if nv is None:
                            should = True
                        else:
                            should = not _prefix_share_root(prefix, nv)
                    if should:
                        marks.add(orphan_row["uid"])
                    continue

                grp.sort(key=lambda x: x[0]["seq"])
                if not any(last == 1 for _, last in grp):
                    continue

                uid_to_inc = {}
                idx = 0
                while idx < len(grp):
                    r, last = grp[idx]
                    if last == 1:
                        members = [(idx, r)]
                        j = idx + 1
                        exp = 2
                        while j < len(grp):
                            rj, lj = grp[j]
                            if lj == exp:
                                members.append((j, rj))
                                exp += 1
                                j += 1
                            else:
                                break
                        if len(members) >= 2:
                            gid = (prefix, idx)
                            for _, m in members:
                                uid_to_inc[m["uid"]] = gid
                        idx = j
                        continue
                    idx += 1

                for k in range(len(grp) - 1, 0, -1):
                    cr, cl = grp[k]
                    pr, pl = grp[k - 1]
                    if cl != pl + 1:
                        cg = uid_to_inc.get(cr["uid"])
                        pg = uid_to_inc.get(pr["uid"])
                        if cg is not None and pg is not None and cg != pg:
                            continue
                        if cg is None and pg is None:
                            chain = 1
                            for fm in range(k + 1, len(grp)):
                                fr, fl = grp[fm]
                                if fl == cl + (fm - k):
                                    chain += 1
                                else:
                                    break
                            if chain >= 2:
                                continue
                        if cl == 1 and cg is not None:
                            continue
                        marks.add(cr["uid"])
        else:
            # 路径B：建堆+堆延长+标记
            if () not in prefix_groups:
                continue
            grp = prefix_groups[()]
            if len(grp) < 2:
                continue
            grp.sort(key=lambda x: x[0]["seq"])
            marks = _path_b_mark(tp, grp, group_data)

        if marks:
            all_marks[tp] = marks

    # ---- Phase 2: seq 固定 + absorbed 链回卷 ----
    changed = False
    absorbed = {}  # absorber_seq → {absorbed_seq, ...}

    for tp in split_type_order:
        marks = all_marks.get(tp, set())
        if not marks:
            continue

        uid_to_seq = {r["uid"]: r["seq"] for r in group_data}
        marked_seqs = sorted(
            [uid_to_seq[uid] for uid in marks if uid in uid_to_seq],
            reverse=True
        )

        for cur_seq in marked_seqs:
            # 找当前行
            cur_idx = next((i for i, r in enumerate(group_data) if r["seq"] == cur_seq), None)
            if cur_idx is None:
                continue
            cur_row = group_data[cur_idx]

            target_seq = cur_row.get("parent_seq", cur_seq - 1)
            if target_seq <= 0:
                continue

            # 找目标行：先查活着的，再查 absorbed 链
            target_idx = next((i for i, r in enumerate(group_data) if r["seq"] == target_seq), None)
            if target_idx is None:
                for absorber_seq, abs_set in absorbed.items():
                    if target_seq in abs_set or target_seq == absorber_seq:
                        target_idx = next((i for i, r in enumerate(group_data) if r["seq"] == absorber_seq), None)
                        if target_idx is not None:
                            break

            if target_idx is not None:
                target_row = group_data[target_idx]
                target_row["content"] = target_row["content"] + cur_row["content"]
                tag = f"回卷{cur_seq}"
                target_row["extra"] = (str(target_row["extra"]) + ";" + tag) if target_row["extra"] else tag

                # 记录 absorbed 链
                tgt_seq = target_row["seq"]
                if tgt_seq not in absorbed:
                    absorbed[tgt_seq] = set()
                absorbed[tgt_seq].add(cur_seq)
                if cur_seq in absorbed:
                    absorbed[tgt_seq].update(absorbed.pop(cur_seq))

                group_data.pop(cur_idx)
                changed = True

    return group_data
# ===== 空前缀过渡切分 =====
def split_empty_prefix_transitions(group_data):
    """
    在所有正常阶段拆分完成后，扫描同一 split_type 内相邻前缀组之间的过渡。
    在 A 前缀组的最后一个块中，寻找 B 前缀组首个差异分量对应的标量序号，
    只切分第一个匹配，作为空前缀边界标记行。
    """
    max_uid = max((r.get("uid", 0) for r in group_data), default=-1)
    uid_counter = max_uid + 1

    splits = []  # (parent_uid, new_row_dict)

    type_rows = {}
    for r in group_data:
        tp = r.get("split_type")
        if tp is not None:
            type_rows.setdefault(tp, []).append(r)

    for tp, rows in type_rows.items():
        if len(rows) < 2:
            continue

        # 在同类型内按序号前缀分子组
        prefix_groups = {}
        for r in rows:
            o = get_ordinal(r["content"])
            if o is None:
                continue
            prefix, last = _ordinal_prefix(o)
            if not prefix:
                continue
            prefix_groups.setdefault(prefix, []).append(r)

        if len(prefix_groups) < 2:
            continue

        sorted_prefixes = sorted(prefix_groups.keys(),
                                 key=lambda p: prefix_groups[p][0]["seq"])

        effective_parent = None
        for i in range(len(sorted_prefixes) - 1):
            a_pref = sorted_prefixes[i]
            b_pref = sorted_prefixes[i + 1]
            used_parent = False

            if len(a_pref) < len(b_pref):
                if a_pref == b_pref[:len(a_pref)]:
                    if effective_parent is None:
                        effective_parent = a_pref
                    continue
            else:
                if effective_parent is not None:
                    a_pref = effective_parent
                    effective_parent = None
                    used_parent = True
                    if len(a_pref) < len(b_pref) and a_pref == b_pref[:len(a_pref)]:
                        effective_parent = a_pref
                        continue

            # 找第一个差异分量
            diff_idx = None
            for j in range(min(len(a_pref), len(b_pref))):
                if a_pref[j] != b_pref[j]:
                    diff_idx = j
                    break
            if diff_idx is None:
                if len(a_pref) > len(b_pref):
                    diff_idx = len(b_pref)
                else:
                    continue

            if diff_idx < len(b_pref):
                target = b_pref[diff_idx]
            else:
                continue

            if used_parent:
                all_in_scope = [r for pkey, rows2 in prefix_groups.items()
                                if len(pkey) >= len(a_pref) and pkey[:len(a_pref)] == a_pref
                                for r in rows2]
                a_rows = sorted(all_in_scope, key=lambda x: x["seq"])
            else:
                a_rows = sorted(prefix_groups[a_pref], key=lambda x: x["seq"])
            last_row = a_rows[-1]

            best_start = None
            best_data = None
            content = last_row["content"]

            search_patterns = list(type_patterns)
            if tp in ("数字点", "数字点点"):
                for p in _TRANSITION_FALLBACK_PATTERNS:
                    if p not in search_patterns:
                        search_patterns.append(p)

            for name, pat, func in search_patterns:
                for m in pat.finditer(content):
                    try:
                        ord_val = func(m)
                    except Exception:
                        continue
                    if isinstance(ord_val, tuple):
                        if diff_idx < len(ord_val):
                            scalar = ord_val[diff_idx]
                        else:
                            continue
                    elif isinstance(ord_val, int):
                        scalar = ord_val
                    else:
                        continue
                    if scalar != target:
                        continue
                    start = m.start()
                    if start > 0 and content[start-1] == '.':
                        if start >= 2 and content[start-2].isdigit():
                            continue
                    pre = content[:start]
                    if pre.strip() == "":
                        continue
                    if best_start is None or start < best_start:
                        best_start = start
                        post = content[start:]
                        best_data = (pre, post, name)

            if best_data is not None:
                pre, post, match_type = best_data
                last_row["content"] = pre
                splits.append((last_row["uid"], {
                    "group": last_row["group"],
                    "content": post,
                    "extra": last_row["extra"],
                    "source_id": last_row.get("source_id", 0),
                    "split_type": match_type,
                    "parent": last_row["uid"],
                }))

    # ===== 正向边界：在每个序号1组的第一个元素的前一个格子里找该章的标量序号 =====
    seq_to_item = {r["seq"]: r for r in group_data}
    for tp, rows in type_rows.items():
        if len(rows) < 1 or tp not in ("数字点", "数字点点"):
            continue
        pg = {}
        for r in rows:
            o = get_ordinal(r["content"])
            if o is None: continue
            prefix, last = _ordinal_prefix(o)
            if not prefix: continue
            pg.setdefault(prefix, []).append(r)
        for prefix, group in pg.items():
            if len(prefix) != 1:
                continue
            ones = [(r, _ordinal_prefix(get_ordinal(r["content"]))[1])
                    for r in group
                    if get_ordinal(r["content"]) is not None
                    and _ordinal_prefix(get_ordinal(r["content"]))[1] == 1]
            if not ones:
                continue
            first_one = min(ones, key=lambda x: x[0]["seq"])[0]
            chapter_num = prefix[0] if prefix else None
            if chapter_num is None:
                continue
            prev_seq = first_one["seq"] - 1
            prev_item = seq_to_item.get(prev_seq)
            if prev_item is None:
                continue
            if chapter_num != 1 and prev_item.get("split_type") is not None:
                continue
            content = prev_item["content"]
            search_pats = list(_TRANSITION_FALLBACK_PATTERNS)
            for p in type_patterns:
                if p not in search_pats:
                    search_pats.append(p)
            best_match = None
            for name, pat, func in search_pats:
                for m in pat.finditer(content):
                    try:
                        ov = func(m)
                    except Exception:
                        continue
                    if isinstance(ov, tuple):
                        scalar = ov[0] if len(ov) > 0 else None
                    elif isinstance(ov, int):
                        scalar = ov
                    else:
                        continue
                    if scalar != chapter_num:
                        continue
                    start = m.start()
                    if start > 0 and content[start-1] == '.' and start >= 2 and content[start-2].isdigit():
                        continue
                    pre = content[:start]
                    if pre.strip() == "":
                        continue
                    best_match = (start, pre, content[start:], name)
            if best_match is not None:
                _, pre, post_match, mtype = best_match
                prev_item["content"] = pre
                splits.append((prev_item["uid"], {
                    "group": prev_item["group"],
                    "content": post_match,
                    "extra": prev_item["extra"],
                    "source_id": prev_item.get("source_id", 0),
                    "split_type": mtype,
                    "parent": prev_item["uid"],
                }))

    # 按 parent 的 seq 倒序插入
    uid_to_seq = {r["uid"]: r["seq"] for r in group_data}
    splits.sort(key=lambda x: uid_to_seq.get(x[0], 0), reverse=True)

    for parent_uid, new_row in splits:
        new_row["uid"] = uid_counter
        uid_counter += 1
        for idx, r in enumerate(group_data):
            if r["uid"] == parent_uid:
                group_data.insert(idx + 1, new_row)
                break

    for idx, r in enumerate(group_data):
        r["seq"] = idx + 1

    return group_data

# ===== 括号二次回卷：主回卷完成后，单独处理括号短堆 =====
def _bracket_secondary_rollback(group_data, group_name, verbose=True):
    """主回卷后，对括号类型做二次回卷。

    建括号堆，对 max_n ≤ 2 且与另一个括号堆 seq 紧邻的短堆，
    标记其全部元素并按回卷逻辑吸收。
    当相连的两个堆都是短堆时，只标记 seq 靠后的那个。
    """
    _v = (lambda *a, **kw: None) if not verbose else print

    # 1. 收集括号行
    bracket_rows = [(r, get_ordinal(r["content"]))
                    for r in group_data if r.get("split_type") == "括号"]
    bracket_rows = [(r, o) for r, o in bracket_rows if o is not None]
    if len(bracket_rows) < 2:
        return group_data
    bracket_rows.sort(key=lambda x: x[0]["seq"])

    # 2. 建堆（同 _path_b_mark 逻辑）
    heaps = []
    current = []
    for r, o in bracket_rows:
        prefix, last = _ordinal_prefix(o)
        if not current:
            current = [(r, prefix, last)]
        elif prefix == current[-1][1] and last == current[-1][2] + 1:
            current.append((r, prefix, last))
        else:
            heaps.append(current)
            current = [(r, prefix, last)]
    if current:
        heaps.append(current)

    heap_info = []
    for h in heaps:
        ordinals = [last for _, _, last in h]
        seqs = [r["seq"] for r, _, _ in h]
        heap_info.append({
            "uids": {r["uid"] for r, _, _ in h},
            "min_seq": min(seqs),
            "max_seq": max(seqs),
            "max_n": ordinals[-1] - ordinals[0] + 1 if ordinals else 0,
            "start_ord": ordinals[0],
            "end_ord": ordinals[-1],
        })

    # 3. 标记符合条件的短堆
    to_rollback = set()
    for i, h in enumerate(heap_info):
        if h["max_n"] > 2:
            continue  # 不是短堆

        connected = False
        for j, other in enumerate(heap_info):
            if i == j:
                continue
            # seq 紧邻判断
            if h["min_seq"] == other["max_seq"] + 1 or h["max_seq"] + 1 == other["min_seq"]:
                connected = True
                # 对方也是短堆 → 只标记 seq 靠后的那个
                if other["max_n"] <= 2:
                    if h["min_seq"] < other["min_seq"]:
                        connected = False  # h 在前，不标记
                break

        if connected:
            to_rollback.update(h["uids"])

    if not to_rollback:
        return group_data

    _v(f"   括号二次回卷：标记 {len(to_rollback)} 个元素")

    # 4. 吸收（调用共享吸收函数）
    uid_to_seq = {r["uid"]: r["seq"] for r in group_data}
    marked_seqs = sorted(
        [uid_to_seq[uid] for uid in to_rollback if uid in uid_to_seq],
        reverse=True
    )
    if marked_seqs:
        group_data = _absorb_marked_seqs(group_data, marked_seqs, "括回")
    return group_data




# ==================== 共享吸收函数 + 路径B二次回卷 ====================


def _absorb_marked_seqs(group_data, marked_seqs, tag_prefix="absorb"):
    """吸收标记的 seq 元素制其 parent_seq 行。

    参数:
        group_data: group data list
        marked_seqs: 要吸收的 seq 列表（需已倒序）
        tag_prefix: extra 中的标记前缀
    返回: 修改后的 group_data
    """
    absorbed = {}
    for cur_seq in marked_seqs:
        cur_idx = next(
            (i for i, r in enumerate(group_data) if r["seq"] == cur_seq), None
        )
        if cur_idx is None:
            continue
        cur_row = group_data[cur_idx]

        target_seq = cur_row.get("parent_seq", cur_seq - 1)
        if target_seq <= 0:
            continue

        target_idx = next(
            (i for i, r in enumerate(group_data) if r["seq"] == target_seq), None
        )
        if target_idx is None:
            for absorber_seq, abs_set in absorbed.items():
                if target_seq in abs_set or target_seq == absorber_seq:
                    target_idx = next(
                        (i for i, r in enumerate(group_data)
                         if r["seq"] == absorber_seq), None
                    )
                    if target_idx is not None:
                        break

        if target_idx is not None:
            target_row = group_data[target_idx]
            target_row["content"] = target_row["content"] + cur_row["content"]
            tag = f"{tag_prefix}{cur_seq}"
            target_row["extra"] = (
                str(target_row["extra"]) + ";" + tag
            ) if target_row["extra"] else tag

            tgt_seq = target_row["seq"]
            if tgt_seq not in absorbed:
                absorbed[tgt_seq] = set()
            absorbed[tgt_seq].add(cur_seq)
            if cur_seq in absorbed:
                absorbed[tgt_seq].update(absorbed.pop(cur_seq))

            group_data.pop(cur_idx)

    return group_data


def _path_b_secondary_rollback(group_data, group_name, split_type_order,
                                 min_stay_length=3, verbose=True):
    """路径B类型二次回卷：对短组（长度<min_stay_length）做回卷。

    在主回卷和括号二次回卷之后运行。
    对每个路径B类型，从 group_data 提取序数序列，
    用 compute_group_marks 计算各组标记，
    将所有 suppressed 组的 uid 加入回卷集。
    """
    _v = (lambda *a, **kw: None) if not verbose else print
    from analyze_split_types import compute_group_marks as _compute_group_marks

    to_rollback = set()
    _PATH_B = PATH_B_TYPES

    for tp in split_type_order:
        if tp not in _PATH_B:
            continue

        rows = [r for r in group_data if r.get("split_type") == tp]
        if len(rows) < 2:
            continue
        rows.sort(key=lambda x: x["seq"])

        # 提取 scalar 序数 + 每个 ordinal → uids 映射
        ords = []
        uid_by_ord = {}
        for r in rows:
            o = get_ordinal(r["content"])
            if o is not None and isinstance(o, int):
                ords.append(o)
                uid_by_ord.setdefault(o, set()).add(r["uid"])

        if not ords:
            continue

        # 用共享函数计算组标记
        marks = _compute_group_marks(ords, min_stay_length)

        for gm in marks:
            if gm["tag"] == "suppressed":
                for o in range(gm["start"], gm["end"] + 1):
                    to_rollback.update(uid_by_ord.get(o, set()))

    if not to_rollback:
        _v(f"    路径B二次回卷：无短组需回卷")
        return group_data

    _v(f"    路径B二次回卷：标记 {len(to_rollback)} 个元素")

    # 吸收
    uid_to_seq = {r["uid"]: r["seq"] for r in group_data}
    marked_seqs = sorted(
        [uid_to_seq[uid] for uid in to_rollback if uid in uid_to_seq],
        reverse=True
    )
    if marked_seqs:
        group_data = _absorb_marked_seqs(group_data, marked_seqs, "短组回卷")

    for idx, r in enumerate(group_data):
        r["seq"] = idx + 1
        r["parent_seq"] = idx

    return group_data

# ===== 核心：分阶段拆分 + 增强回卷 =====
def split_single_group_with_rollback(group_data, group_name, split_types_override=None, verbose=True, score_collector=None):
    _v = (lambda *a, **kw: None) if not verbose else print
    uid_counter = 0
    for item in group_data:
        item["uid"] = uid_counter
        item["parent"] = None
        item["split_type"] = None
        uid_counter += 1

    group_data.sort(key=lambda x: int(x["seq"]) if isinstance(x["seq"], (int, float)) else 0)
    for idx, item in enumerate(group_data): item["seq"] = idx + 1

    # ---- 分阶段拆分 ----
    split_type_list = split_types_override if split_types_override is not None else SPLIT_TYPES
    for stage_type in split_type_list:
        stage_types = {stage_type}
        _v(f"  [{group_name}] 开始阶段拆分：类型 [{stage_type}]")
        i = 0
        while i < len(group_data):
            cur = group_data[i]
            res = find_split_point_for_types(cur["content"], stage_types)
            if res:
                pre, post, split_type = res
                try:
                    _v(f"    seq{cur['seq']} 拆出 [{cur['content'][:30]}...] -> [{pre[:20]}...] + [{post[:20]}...]")
                except UnicodeEncodeError:
                    _v(f"    seq{cur['seq']} 拆出 (编码特殊字符, 略)")
                cur["content"] = pre
                new_item = {
                    "group": group_name, "seq": cur["seq"] + 1,
                    "content": post, "extra": cur["extra"],
                    "source_id": cur.get("source_id", cur["uid"]),
                    "uid": uid_counter, "parent": cur["uid"],
                    "split_type": split_type,
                }
                uid_counter += 1
                group_data.insert(i + 1, new_item)
                for idx, r in enumerate(group_data): r["seq"] = idx + 1
            i += 1
        _v(f"    阶段 [{stage_type}] 拆分结束，当前共 {len(group_data)} 行。")

    # ---- 空前缀过渡切分 ----
    _v(f"  [{group_name}] 开始空前缀过渡切分...")
    group_data = split_empty_prefix_transitions(group_data)
    _v(f"    过渡切分结束，当前共 {len(group_data)} 行。")

    # ---- 统一固定 seq + parent_seq（所有拆分完毕，物理顺序永不变） ----
    for idx, r in enumerate(group_data):
        r["seq"] = idx + 1
        r["parent_seq"] = idx  # idx=0 → parent_seq=0 表示无父行

    # ---- 全局后向回卷 ----
    _v(f"  [{group_name}] 开始全局后向回卷...")
    group_data = global_backward_rollback(group_data, group_name, split_type_list, score_collector)
    _v(f"    回卷结束，当前共 {len(group_data)} 行。")

    # ---- 括号类型二次回卷 ----
    if "括号" in split_type_list:
        _v(f"  [{group_name}] 开始括号二次回卷...")
        group_data = _bracket_secondary_rollback(group_data, group_name, verbose=verbose)
        for idx, r in enumerate(group_data):
            r["seq"] = idx + 1
            r["parent_seq"] = idx
        _v(f"    括号二次回卷结束，最终共 {len(group_data)} 行。")

    # ---- 路径B类型二次回卷（组长度判据） ----
    _v(f"  [{group_name}] 开始路径B二次回卷...")
    group_data = _path_b_secondary_rollback(group_data, group_name, split_type_list,
                                              min_stay_length=3, verbose=verbose)
    for idx, r in enumerate(group_data):
        r["seq"] = idx + 1
        r["parent_seq"] = idx
    _v(f"    路径B二次回卷结束，最终共 {len(group_data)} 行。")
    return group_data

# ===== 数据库支持 =====
def get_law_text_from_db(law_id):
    try:
        import pymysql
    except ImportError:
        print("pymysql未安装，请执行 pip install pymysql")
        sys.exit(1)
    conn = pymysql.connect(**DB_CONFIG)
    try:
        with conn.cursor() as cursor:
            cursor.execute("SELECT attachment_url FROM mt_kb_law_metadata WHERE law_id = %s", (law_id,))
            res = cursor.fetchone()
            return res[0] if res else ""
    finally:
        conn.close()

def clean_html(raw_text):
    """数据库原文清洗：去标签、解常见实体、压缩空白。"""
    txt = re.sub(r"<[^>]+>", "", raw_text)
    txt = (txt.replace("&nbsp;", " ")
              .replace("&amp;", "&")
              .replace("&lt;", "<")
              .replace("&gt;", ">"))
    txt = re.sub(r"\s+", " ", txt).strip()
    return txt

def process_raw_text_to_single_row(raw_text, law_id):
    """直接传入原始文本，不做清洗"""
    return [{
        "group": law_id, "seq": 1,
        "content": raw_text, "extra": None,
        "source_id": 0, "split_type": None
    }]

# ===== 主流程 =====
def parse_and_reorder(text, split_type="条"):
    """全局排序纠错：对纯文本做全文匹配 → 切段 → 建连续性堆 → 递增组链接 → 游离堆回卷。"""
    tp_entry = None
    for name, pat, func in type_patterns:
        if name == split_type:
            tp_entry = (pat, func)
            break
    if tp_entry is None:
        raise ValueError(f"parse_and_reorder: split_type '{split_type}' 不在 type_patterns 中")

    pattern, ord_func = tp_entry

    matches = list(pattern.finditer(text))
    if not matches:
        return []

    records = []
    for idx, m in enumerate(matches, start=1):
        article_no = ord_func(m)
        if not isinstance(article_no, int):
            raise ValueError(
                f"parse_and_reorder 仅支持 scalar 序数类型（int），"
                f"split_type='{split_type}' 返回了 {type(article_no).__name__}: {article_no}"
            )
        records.append({"order": idx, "article_no": article_no, "segment_idx": None})

    if not records:
        return []

    starts = [m.start() for m in matches]
    paragraphs = [text[:starts[0]]]
    for i in range(len(matches)):
        start = starts[i]
        end = starts[i + 1] if i < len(matches) - 1 else len(text)
        paragraphs.append(text[start:end])

    for i, rec in enumerate(records, start=1):
        rec["segment_idx"] = i

    heaps = []
    current = None
    for rec in records:
        if current is None:
            current = {
                "heap_no": None,
                "entries": [rec],
                "start_article": rec["article_no"],
                "end_article": rec["article_no"],
            }
        elif rec["article_no"] == current["entries"][-1]["article_no"] + 1:
            current["entries"].append(rec)
            current["end_article"] = rec["article_no"]
        else:
            heaps.append(current)
            current = {
                "heap_no": None,
                "entries": [rec],
                "start_article": rec["article_no"],
                "end_article": rec["article_no"],
            }
    if current:
        heaps.append(current)

    for i, h in enumerate(heaps, start=1):
        h["heap_no"] = i
    heap_map = {h["heap_no"]: h for h in heaps}

    unassigned = {h["heap_no"] for h in heaps}
    inc_groups = []
    group_no = 0
    search_start = 1
    while search_start in unassigned or any(h_no >= search_start for h_no in unassigned):
        candidates = [h_no for h_no in unassigned if h_no >= search_start]
        if not candidates:
            break
        start_h_no = min(candidates)
        cur_h = heap_map[start_h_no]
        group = [cur_h]
        unassigned.remove(start_h_no)

        while True:
            next_h_no = None
            for h_no in sorted(unassigned):
                if heap_map[h_no]["start_article"] == cur_h["end_article"] + 1:
                    next_h_no = h_no
                    break
            if next_h_no is None:
                break
            next_h = heap_map[next_h_no]
            unassigned.remove(next_h_no)
            group.append(next_h)
            cur_h = next_h

        group_no += 1
        inc_groups.append({"group_no": group_no, "heaps": group})
        search_start = group[-1]["heap_no"] + 1

    tag_r = set()
    for group_info in inc_groups:
        heaps_in_group = group_info["heaps"]
        start_h = heaps_in_group[0]["heap_no"]
        end_h = heaps_in_group[-1]["heap_no"]
        group_set = {h["heap_no"] for h in heaps_in_group}
        for h_no in range(start_h, end_h + 1):
            if h_no in heap_map and h_no not in group_set:
                tag_r.add(h_no)

    for group_info in reversed(inc_groups):
        first_rec = group_info["heaps"][0]["entries"][0]
        if first_rec["article_no"] != 1:
            for h in group_info["heaps"]:
                tag_r.add(h["heap_no"])

    rollback_list = sorted(tag_r)
    while rollback_list:
        h_no = rollback_list.pop()
        target = heap_map[h_no]
        while target["entries"]:
            entry = target["entries"].pop()
            seg = entry["segment_idx"]
            if seg > 0:
                paragraphs[seg - 1] += paragraphs[seg]
                paragraphs[seg] = ""
        target["start_article"] = None
        target["end_article"] = None

    reordered = []
    for group_info in inc_groups:
        for h in group_info["heaps"]:
            for entry in h["entries"]:
                seg = entry["segment_idx"]
                content = paragraphs[seg]
                if content:
                    reordered.append(content)

    prefix = paragraphs[0].strip() if paragraphs and paragraphs[0] else ""
    if prefix:
        reordered.insert(0, paragraphs[0])

    return reordered


def main():
    wb = openpyxl.load_workbook(INPUT_FILE)
    ws = wb[SHEET_NAME]

    all_rows = []
    for row in ws.iter_rows(min_row=HEADER_ROW + 1, max_row=ws.max_row, values_only=True):
        if row[0] is None: continue
        group = str(row[0]).strip()
        seq = row[1]
        content = row[2] if len(row) > 2 else ""
        extra = row[3] if len(row) > 3 else None
        all_rows.append({"group": group, "seq": seq,
                         "content": str(content) if content else "",
                         "extra": extra})
    for sid, r in enumerate(all_rows): r["source_id"] = sid

    print(f"共读取 {len(all_rows)} 行数据。")

    if USE_DB:
        target_groups = [str(g).strip() for g in LAW_IDS if str(g).strip()]
    else:
        target_groups = [str(g).strip() for g in LAW_IDS] if LAW_IDS else []

    unprocessed = [r for r in all_rows if r["group"] not in target_groups]
    final_rows = unprocessed.copy()

    for gn in target_groups:
        if USE_DB:
            print(f"\n处理 {gn} ...")
            raw_text = get_law_text_from_db(gn)
            if not raw_text:
                print(f"  未获取到文本 {gn}")
                continue

            clean_text = clean_html(raw_text)

            # ---- 纯文本段落拆分：走独立逻辑 ----
            if "纯文本段落拆分" in SPLIT_TYPES:
                paragraphs = split_plain_by_paragraphs(clean_text)
                gdata = []
                for idx, content in enumerate(paragraphs, start=1):
                    gdata.append({
                        "group": gn, "seq": idx,
                        "content": content, "extra": None,
                        "source_id": 0, "split_type": None
                    })
                print(f"  纯文本段落拆分 → {len(gdata)} 段")

                other_types = [t for t in SPLIT_TYPES if t != "纯文本段落拆分"]
                if other_types:
                    print(f"  后续拆分类型: {other_types}")
                    gdata = split_single_group_with_rollback(gdata, gn, split_types_override=other_types)
                    if gdata is None:
                        print("  处理结果 None, 跳过")
                        continue
                final_rows.extend(gdata)
                continue

            reordered = parse_and_reorder(clean_text, DB_PRIMARY_TYPE)
            if not reordered:
                reordered = [clean_text]

            gdata = []
            for idx, content in enumerate(reordered, start=1):
                gdata.append({
                    "group": gn, "seq": idx,
                    "content": content, "extra": None,
                    "source_id": 0, "split_type": None
                })
            print(f"  parse_and_reorder({DB_PRIMARY_TYPE}) → {len(gdata)} 段")

            secondary = DB_SECONDARY_TYPES
            if secondary is None:
                secondary = [tp for tp in SPLIT_TYPES if tp != DB_PRIMARY_TYPE]
            if secondary:
                print(f"  DB 次要拆分类型: {secondary}")
                gdata = split_single_group_with_rollback(gdata, gn, split_types_override=secondary)
                if gdata is None:
                    print("  处理结果 None, 跳过")
                    continue
        else:
            gdata = [r for r in all_rows if r["group"] == gn]
            if not gdata:
                print(f"未找到组 '{gn}'，跳过")
                continue

            gdata = split_single_group_with_rollback(gdata, gn)
            if gdata is None:
                print("处理结果 None, 跳过")
                continue

        final_rows.extend(gdata)

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = SHEET_NAME
    headers = ["组", "序号", "内容", "保留列"]
    for col, h in enumerate(headers, start=1):
        ws_out.cell(row=1, column=col, value=h)
    for i, row in enumerate(final_rows, start=2):
        ws_out.cell(row=i, column=1, value=row["group"])
        ws_out.cell(row=i, column=2, value=row["seq"])
        ws_out.cell(row=i, column=3, value=row["content"])
        ws_out.cell(row=i, column=4, value=row["extra"])

    wb_out.save(OUTPUT_FILE)
    print(f"\n全部完成！结果保存至 {OUTPUT_FILE}")

if __name__ == "__main__":
    main()